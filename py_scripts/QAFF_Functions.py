import os
import sys
import numpy as np
from File_Functions import dictFromFile



class Dashboard():
    
    def __init__(self):
        cwd = os.getcwd()
        self.cwd = cwd
        self.basedir = 'U:\QAFF'
        self.dbdir = '%s\Dropbox\Unknome\Databases\QAFFDB' %self.cwd
        self.QAFFdir = '%s\Dropbox\Unknome\Screens\QAFF' %self.cwd
        self.workdir = os.path.join(self.QAFFdir, 'iQAFF') 
        self.gridmapdir = os.path.join(self.QAFFdir, 'GridmapFiles')
        self.platesdir = os.path.join(self.workdir, 'Plates_data')
        self.depositsdir = os.path.join(self.workdir, 'Deposits_data')
        self.dheads = ['plateID', 'depositID', 'includeDeposit', 'Xposition', 'Yposition', 'area', 'perimeter', 'circularity', 'ROD', 'IOD', 'meanB', 'meanG', 'meanR', 'meanH', 'meanL', 'meanS']
        self.pickledir = os.path.join(self.workdir, 'PickleFiles')
        self.batchlist = self.fetchBatchlist()
        self.batchnumbers = self.fetchBatchnumbers()
        self.ctrlnames = self.controlsNames()
        self.ctrlnamesList = self.controlsNamesList()
        QAFFdirs = [self.dbdir, self.gridmapdir, self.depositsdir, self.platesdir]
        for dirpath in QAFFdirs:
            if not os.path.exists(dirpath):
                os.makedirs(dirpath)
                
    def getMembers(self):
        import inspect
        methods = inspect.getmembers(self, predicate = inspect.ismethod)
        print(methods)
        return
    
    def fetchBatchlist(self):
        batchlist = [os.path.splitext(name)[0] for name in os.listdir(self.dbdir) if name.startswith('QAFF')]
        batchlist = sorted(batchlist, key = lambda x:int(x[4:]))
        return batchlist
        
    def fetchBatchnumbers(self):
        batchlist = self.fetchBatchlist()
        batchnumbers = [int(val[4:]) for val in batchlist]
        return batchnumbers 
        
    def controlsDict(self):
        #Define dictionary
        controlsDict = {'EMPTY': 'Empty', 'Emp': 'Empty', 'EMP': 'Empty', 'GFPI': 'GFPi', 'W1118': 'w1118', 'Lk' : 'LK', 'lk' : 'LK'}
        return controlsDict 
                     
    def controlsNames(self):
        from collections import Counter as counter
        counterDict = counter(self.controlsDict().values())
        ctrlNames = counterDict.keys()
        return ctrlNames
    
    def controlsNamesList(self):
        ctrlNameslist = self.controlsDict().keys()
        return ctrlNameslist  
          
    def resetWorkEnv(self, batch = 'all'):
        '''It takes a batch number, copies its database file to the database directory and exports the deposits and plates tables
        to the respective data directories. If batch = 'all' it applies these operations to all batches in the batchlist. Irrespective of the keyword
        arg value it rebuilds the gridmap and areas databases and dictionary objects.'''
        QAFFDB().copyToDBdir(batch = batch)
        QAFFDB().exportTableData(batch = batch)
        print('\nBuilding SQLKeys dictionary.')
        QAFFObjects().buildSQLkeysDict()
        print('\nBuilding gridmap database.')
        QAFFDB().buildGridmapDB()
        print('\nBuilding areas database.')
        QAFFDB().buildAreasDB()
        print('\nBuilding platesmap database.')
        QAFFDB().buildPlatesMapDB()
        print('\nBuilding ROD colour filter.')
        QAFFObjects().buildRODColourFilter()
        print('\nBuilding ROD size filter.')
        QAFFObjects().buildRODSizeFilter()
        print('\nMapping dataset keys.')
        QAFFObjects().dsetKeysMapper()
        print('\nCalculating QAFF metrics for the dataset.')
        DataOperations().calculateMetricsForDataset()
        print('\nBuilding the dataset metrics dictionary.')
        DataOperations().buildDatasetMetricsObj()
        print('\nBuilding typeIRODkeys dictionary.')
        QAFFObjects().buildTypeIRODkeysDict()
        print('\nFetching area and void mismatches.')
        areaMismatches, voidMismatches = DataOperations().fetchGridmapMismatches()
        print('\nPurging mismatches and pickle directories.')
        DataOperations().purgePickleDir(fnum = 2)
        DataOperations().purgeMismatchesDir(fnum = 2)
        return



class QAFFDB(Dashboard):
    
    def __init__(self):
        Dashboard.__init__(self)
                                  
    def copyToDBdir(self, batch = 'all'):
        '''It takes a batch number and copies its database file to the database directory. If batch = 'all' it copies all database files 
        for all batches in the batchlist that have a valid database file in basedir.'''
        from File_Functions import listPartition
        import shutil
        #test whether batch exists in the batchlist
        if batch == 'all':
            batchlist = self.batchlist
        else:
          if isinstance(batch, int):
            batch = [batch]
          batchlist = ['QAFF%s' %val for val in batch]
          batchlist, mismatches = listPartition(lambda x:x in self.batchlist, batchlist)
          if len(mismatches)>0:
              mismatches_str = (',').join(mismatches)
              print('Batches %s are not in the batchlist\n' %mismatches_str)
        #copy database files to database directory
        for name in batchlist:
          dirpath = os.path.join(self.basedir, '%s_scans' %name) 
          dbpath = os.path.join(dirpath, '%s.sqlite' %name)
          outpath = os.path.join(self.dbdir, '%s.sqlite' %name)
          if os.path.exists(dbpath):
              print('Copying file %s.sqlite' %name)
              shutil.copy(dbpath, outpath)
          else:
              print('The file %s.sqlite does not exist.\n' %name)
        return
        
        
    def exportTableData(self, batch = 'all'):
        from File_Functions import listPartition
        import sqlite3
        '''Given a batch number (int) or a sequence of batch numbers (list, tuple) it exports the deposits and plates tables 
        as text files for each batch; if batch = 'all' it does so for all batches for which there is a valid datatbase file in the database directory (dbdir).'''
        #database file paths
        if batch == 'all':
            dbfilepaths = [os.path.join(self.dbdir, '%s.sqlite' %name) for name in self.batchlist]
        else:
            if isinstance(batch, int):
                batch = [batch]
            batchlist = ['QAFF%s' %val for val in batch]
            batchlist, mismatches = listPartition(lambda x:x in self.batchlist, batchlist)
            if len(mismatches)>0:
                mismatches_str = (',').join(mismatches)
                print('Batches %s are not in the batchlist\n' %mismatches_str)
            dbfilepaths = [os.path.join(self.dbdir, '%s.sqlite' %name) for name in batchlist]#fetch file paths 
                
        for i, filepath in enumerate(dbfilepaths):
            #Connect to database
            if os.path.exists(filepath):
                db = sqlite3.connect(filepath)
                cursor = db.cursor()
            else:
                print('File %s.sqlite does not exist.\n' %self.batchlist[i])
                continue
            #fetch data from database tables
            tablenames = ['deposits', 'plates']
            tablesdir = [self.depositsdir, self.platesdir]
            for j, tablename in enumerate(tablenames):
                cursor.execute('''SELECT * FROM %s ''' %tablename)
                tabledata =  cursor.fetchall()
                if batch == 'all':
                    print('Exporting %s table for %s' %(tablename, self.batchlist[i]))
                #reformat rows
                rows = [('\t').join(map(str, tupl)) for tupl in tabledata]
                rows = [row + '\n' for row in rows]
                #fetch headings
                headings = list(map(lambda x: x[0], cursor.description))
                headings = ('\t').join(headings)#reformat
                #write data to file
                txtfilepath = os.path.join(tablesdir[j], '%s_%s.txt' %(self.batchlist[i], tablename))
                with open(txtfilepath, 'w') as f:
                    f.write(headings + '\n')
                    f.writelines(rows)
        return 

    
    def buildGridmapDB(self):
        from datetime import datetime
        import cPickle as pickle
        from itertools import chain
        import sqlite3
        '''It creates a gridmap database from the latest gridmap object and serialises it. '''
        #connect to database
        dbpath = os.path.join(self.dbdir, 'Gridmap.sqlite')
        db = sqlite3.connect(dbpath)
        cursor = db.cursor() 
        #load gridmap object
        print('\nBuilding gridmap dictionary object.')
        QAFFObjects().buildGridmapObject()
        gridmap = QAFFObjects().loadGridmap()
        #fetch existing tablenames
        tablenames = list(cursor.execute(''' SELECT NAME FROM sqlite_master WHERE TYPE = "table"'''))
        tablenames = list(chain(*tablenames))
        #create database
        for batch in gridmap.keys():
            #fetch rows
            rows = gridmap[batch]
            if batch in tablenames:#test whether table already exists
                cursor.execute('''DROP TABLE IF EXISTS %s''' %batch)
                db.commit()
            #create table and insert values
            print('Creating gridmap table %s' %batch)    
            cursor.execute('''CREATE TABLE %s (qKey TEXT NOT NULL, StockID TEXT NOT NULL, Flycount INT NOT NULL)''' %batch)#create table
            insertStatement = '''INSERT INTO %s (qKey, StockID, Flycount) VALUES (?,?,?)''' %batch #insert data
            cursor.executemany(insertStatement, rows)
            db.commit()
        db.close()
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')   
        #Serialize gridmap
        path = os.path.join(self.pickledir, 'gridMap_%s.pickle' %time)
        with open(path, 'wb') as f:
            pickle.dump(gridmap, f, protocol = 2)
        return
        
        
    def buildAreasDB(self):
        import cPickle as pickle
        from datetime import datetime
        from itertools import chain
        import sqlite3
        '''It creates an areas database from the latest areas object and serialises it. ''' 
        #connect to database
        dbpath = os.path.join(self.dbdir, 'AreasDB.sqlite')
        db = sqlite3.connect(dbpath)
        cursor = db.cursor() 
        #load areas object
        QAFFObjects().buildAreasObject()
        areasDict = QAFFObjects().loadAreasDict()
        #fetch existing tablenames
        tablenames = list(cursor.execute(''' SELECT NAME FROM sqlite_master WHERE TYPE = "table"'''))
        tablenames = list(chain(*tablenames))
        #create database
        for batch in areasDict.keys():
            #fetch and reformat rows
            rows = areasDict[batch].items()
            rows = [(scanID, float(area)) for (scanID, area) in rows]
            if batch in tablenames:#test whether table already exists
                cursor.execute('''DROP TABLE IF EXISTS %s''' %batch)
                db.commit()
            #create table and insert values
            print('Creating areas table %s' %batch)    
            cursor.execute('''CREATE TABLE %s (File TEXT NOT NULL, Area REAL NOT NULL)''' %batch)#create table
            insertStatement = '''INSERT INTO %s (File, Area) VALUES (?,?)''' %batch #insert data
            cursor.executemany(insertStatement, rows)    
            db.commit()
        db.close()
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')   
        #Serialize gridmap
        path = os.path.join(self.pickledir, 'areasDict_%s.pickle' %time)
        with open(path, 'wb') as f:
            pickle.dump(areasDict, f, protocol = 2)
        return  
         
    def buildPlatesMapDB(self):
        from itertools import chain
        from datetime import datetime
        import cPickle as pickle
        import sqlite3
        #build platesMap object
        QAFFObjects().buildPlatesMapObject()
        platesMap = QAFFObjects().loadPlatesMap()
        platesMap_keys = ['qkey', 'stockId', 'flycount', 'plate area']
        #Fetch rows from platesMap
        rows = [([sqlkey.split('_')[0]], [sqlkey.split('_')[1]], [platesMap_val[key] for key in platesMap_keys]) for (sqlkey, platesMap_val) in platesMap.items()]
        rows = [tuple(chain(*tupl)) for tupl in rows]
        rows = sorted(rows, key = lambda x:int(x[0]))# sort rows according to batch number
        rows = [(int(a), int(b), c, d, int(e), float(f)) for (a,b,c,d,e,f) in rows]
        #connect to database
        dbpath = os.path.join(self.dbdir, 'PlatesMapDB.sqlite')
        db = sqlite3.connect(dbpath)
        cursor = db.cursor() 
        #fetch existing tablenames
        tablenames = list(cursor.execute(''' SELECT NAME FROM sqlite_master WHERE TYPE = "table"'''))
        tablenames = list(chain(*tablenames))
        #test whether table already exists
        if 'PlatesMap' in tablenames:
            cursor.execute('''DROP TABLE IF EXISTS PlatesMap''' )
            db.commit()
        #create table and insert values    
        cursor.execute('''CREATE TABLE PlatesMap (Batch INT NOT NULL, SQLkey INT NOT NULL, Qkey TEXT NOT NULL, StockID TEXT NOT NULL, Flycount INT NOT NULL, PlateArea REAL NOT NULL)''')#create table
        insertStatement = '''INSERT INTO PlatesMap (Batch, SQLkey, QKey, StockID, Flycount, PlateArea) VALUES (?,?,?,?,?,?)''' #insert data
        cursor.executemany(insertStatement, rows)    
        db.commit()
        db.close()
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')  
        #Serialize platesMapDict
        path = os.path.join(self.pickledir, 'platesMap_%s.pickle' %time)
        with open(path, 'wb') as f:
            pickle.dump(platesMap, f, protocol = 2)
        return
        
    def excludeTypeIRODs(self, databasedir):
        import sqlite3
        #load data objects
        type1RODkeys = QAFFObjects().loadTypeIRODkeys()
        #update includeDeposit column in each batch database
        for batch in type1RODkeys.keys():
            print('Updating QAFF%s.sqlite' %batch)
            dbpath = os.path.join(databasedir, 'QAFF%s.sqlite' %batch)
            db = sqlite3.connect(dbpath)
            cursor = db.cursor()
            keyset = type1RODkeys[batch]
            for key in keyset:
                depositID = int(key.split('_')[0])
                #update data
                updateStatement = '''UPDATE deposits SET includeDeposit = ? WHERE id = ?'''
                cursor.execute(updateStatement, (0,depositID))
            db.commit()
            db.close()
        return

              
                                           
class QAFFObjects(Dashboard):
    
    def __init__(self):
        Dashboard.__init__(self)
        if not os.path.exists(self.pickledir):
            os.mkdirs(self.pickledir)
        self.stocksdir = os.path.join(self.pickledir, 'Stocks')
        
    def getMembers(self):
        import inspect
        methods = inspect.getmembers(self, predicate = inspect.ismethod)
        print(methods)
        return
          
            
    def loadWorkDataset(self):
        #load platesMap dictionary
        platesmapDict = self.loadPlatesMap()
        #fetch stockIDs
        platesmap = platesmapDict.values()
        workdset = [platemap['stockId'][:-2] for platemap in platesmap]
        workdset = [int(Id) for Id in workdset if Id.isnumeric()]#filter out controls
        workdset = sorted(list(set(workdset))); workdset = ['JS%s' %Id for Id in workdset]#reformat list
        workdset = workdset + self.ctrlnames #add standard controls names 
        return workdset
        
                        
    def buildSQLkeysDict(self, batch = 'all'):
        import cPickle as pickle
        from datetime import datetime
        '''It takes a batchnumber (int) or a sequence of batchnumbers (tuple, list, ndarray) and returns a serialized dictionary
        {'sqlkey': 'scanId'} for each batch. If batch = 'all' or a sequence returns a nested dictionary: {'batchname':{'sqlkey': 'scanId'}}.''' 
        if batch == 'all':
            numbatches = len(os.listdir(self.platesdir))
            batchnames = ['QAFF%i' %(idx+1) for idx in xrange(numbatches)]
            platefiles = ['QAFF%i_plates.txt' %(idx+1) for idx in xrange(numbatches)]
            platepath = [os.path.join(self.platesdir, platefile) for platefile in platefiles]
        elif isinstance(batch, (tuple, list, np.ndarray)):
            batch.sort()
            batchnames = ['QAFF%i' %j for j in batch]
            platefiles = ['QAFF%i_plates.txt' %item for item in batch]
            platepath = [os.path.join(self.platesdir, platefile) for platefile in platefiles]
        elif isinstance(batch, int):
            platefiles = 'QAFF%i_plates.txt' %batch
            platepath = [os.path.join(self.platesdir, platefiles)]
        #build dictionary       
        sqlkeysDict_inv = [dictFromFile(path, 0, usecols = 1, colasarr = False) for path in platepath]
        sqlkeysDict = [dict(zip(item.values(), item.keys())) for item in sqlkeysDict_inv]
        if batch == 'all' or isinstance(batch, (tuple, list, np.ndarray)):
            sqlkeysDict = dict(zip(batchnames, sqlkeysDict))
            #Fetch current time and date
            time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            #Serialize sqlkeysDict
            picklefile = os.path.join(self.pickledir,'sqlkeysDict_%s.pickle' %time)
            with open(picklefile, 'wb') as f:
                pickle.dump(sqlkeysDict, f, protocol = 2)
        return


    def loadSQLkeysDict(self):
        import cPickle as pickle
        '''Loads the latest version of SQLkeyDict.'''
        #Load sqlkeysDict
        filenames = [item for item in os.listdir(self.pickledir) if item.startswith('sqlkeysDict')]
        filenames.sort()
        picklefile = filenames[-1]  
        picklePath = os.path.join(self.pickledir, picklefile)
        with open(picklePath, 'rb') as f:
                sqlkeysDict = pickle.load(f)      
        return sqlkeysDict
    
        
    def buildAreasObject(self):
        from datetime import datetime
        import cPickle as pickle
        '''It returns a nested dictionary: {'batchname':{'scanId': 'plate area'}}.'''
        #fetch paths of areas files
        areasPathlist = ['%s\QAFF%i_scans\Areas\QAFF%i_AreasResults.txt' %(self.basedir, number, number) for number in self.batchnumbers]
        #build areas object       
        areasDict = {} 
        for i, path in enumerate(areasPathlist):
            try:
                assert os.path.exists(path), 'File %s does not exist. \n' % os.path.split(path)[1]
                batchDict = dictFromFile(path, 0, usecols = 3, colasarr = False)
                areasDict[self.batchlist[i]] = batchDict
            except AssertionError as ae: 
                print(ae)
        #fetch current time
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #serialise dictionary
        filepath = os.path.join(self.pickledir, 'areasDict_%s.pickle' %time)
        with open(filepath, 'wb') as f:
            pickle.dump(areasDict, f, protocol = 2)          
        return
        

    def loadAreasDict(self):
        import cPickle as pickle
        '''Loads the latest version of areasDict.'''
        #Load areasDict
        filenames = [item for item in os.listdir(self.pickledir) if item.startswith('areasDict')]
        filenames.sort()
        picklefile = filenames[-1]  
        picklePath = os.path.join(self.pickledir, picklefile)
        with open(picklePath, 'rb') as f:
                areasDict = pickle.load(f)
        return areasDict

    
    def  fetchFlyNumber(self, cell):
            '''It takes a cell's identifier from a workbook and extracts from that cell's comments the respective flynumber.
            It returns the flynumber(int)'''
            comment = [item for item in cell.comment.text.split() if item.isdigit()]
            if 0<len(comment)<2:
                flynumber = int(comment[0]) 
            return flynumber

    
    def buildGridmapObject(self, lastcell=100):
        from datetime import datetime
        sys.path.append('C:\Python27\Lib\site-packages')
        from openpyxl import load_workbook
        import cPickle as pickle  
        '''It reads the gridmap excel file and builds a serialised dictionary from it: {'batchname':[qkey, stockId, flycount]}. Optional
        argument lastcell sets the number of cells that are read from each worksheet.'''
        #load latest gridmap file
        gridmapFiles = [(filename, int(filename.split('_')[1][3:9])) for filename in os.listdir(self.gridmapdir) if filename.startswith('QAFFscreen')]
        gridmapFiles = sorted(gridmapFiles, key = lambda x:x[1])
        gridpath = os.path.join(self.gridmapdir, gridmapFiles[-1][0])
        wb = load_workbook(gridpath)#Load workbook
        #Load worksheets
        sheetNames = [name for name in wb.get_sheet_names() if name.startswith('QAFF')]#fetch sheet titles
        wslist = [wb[name] for name in sheetNames]
        #Fetch cells 
        cells_byCols = [zip(*[cell for cell in ws['C2:E%i' %lastcell]]) for i, ws in enumerate(wslist)]
        #Define scans positional grid keys: Qkeys
        rowlabels = ['A', 'B', 'C', 'D']
        scangrid = [['S%i%s' %(j+1,label) for label in rowlabels] for j in xrange(lastcell/4)]
        scangrid = [item for lst in scangrid for item in lst]#upack items
        scangrid = [['%s%i' %(item, i+1) for item in scangrid] for i in xrange(3)]#add column number
        #Fetch cells' values
        cellVals_byCols = [[[cell.value for cell in col] for col in ws] for ws in cells_byCols]
        #Fetch flynumber from comments       
        flynumbers = [[[8 if isinstance(cell.comment, type(None)) else self.fetchFlyNumber(cell) for cell in col] for col in ws] for ws in cells_byCols]
        #Assemble gridmap
        gridmap = [[zip(scangrid[j], col, flynumbers[i][j]) for j, col in enumerate(ws)] for i, ws in enumerate(cellVals_byCols)]
        gridmap = [zip(*ws) for ws in gridmap] # order cells by rows
        gridmap = [[cell for row in ws for cell in row] for ws in gridmap]#unpack rows  
        #Purge gridmap of empty cells
        gridmap = [[cell for cell in ws if not isinstance(cell[1], type(None))] for ws in gridmap]
        #Build gridmap dictionary
        gridmapDict = dict(zip(sheetNames, gridmap))
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #serialise dictionary
        picklepath = os.path.join(self.pickledir, 'gridMap_%s.pickle' %time) 
        with open(picklepath, 'wb') as f:
            pickle.dump(gridmapDict, f, protocol = 2)   
        return    
    

    def loadGridmap(self):
        import cPickle as pickle
        '''Loads the latest version of gridmap dictionary.'''
        #Load the grid map
        filenames = [item for item in os.listdir(self.pickledir) if item.startswith('gridMap')]
        filenames.sort()
        picklefile = filenames[-1]  
        picklePath = os.path.join(self.pickledir, picklefile)
        with open(picklePath, 'rb') as f:
                gridmapDict = pickle.load(f)
        return gridmapDict
        
        
    def buildPlatesMapObject(self):
        from datetime import datetime
        import cPickle as pickle
        from itertools import ifilter
        '''It builds a plate dictionary:  {key: {qkey, stockId, flycount, plate area}} and serialises it; key = batchnumber_sqlKey.'''
        #Load gridmap dictionary
        gridmapDict = self.loadGridmap()
        #Define batchnames
        batchnames = gridmapDict.keys()
        batchnames = sorted(batchnames, key = lambda x:int(x[4:]))#sort batchnames
        #Fetch ordered gridmap values
        gridmap = [gridmapDict[name] for name in batchnames]
        #Load areas dictionary
        sqlKeysDict = self.loadSQLkeysDict()
        areasDict = self.loadAreasDict()
        #Load mismatches        
        areaMismatches, voidMismatches = DataOperations().fetchGridmapMismatches()
        mismatches = list(set(areaMismatches) | set(voidMismatches))
        #build platesmap object
        platesMap = {}
        mapkeys = ['qkey', 'stockId', 'flycount', 'plate area']
        for i, sublist in enumerate(gridmap):
            print('\nBuilding %s plate map. ' %batchnames[i])
            iterfilter = ifilter(lambda x: '%i%s' %(i+1, x[0]) not in mismatches, sublist)#filter out mismatches
            for (qkey, stockId, flynumb) in iterfilter:
                pmapKey = '%i_%s' %(i+1, sqlKeysDict[batchnames[i]][DataOperations().fromQkeyToScan(qkey)[0]])#fecth sqlKey
                plateArea = areasDict[batchnames[i]][DataOperations().fromQkeyToScan(qkey)[0]]#fetch plate area
                plate = ('%i%s' %(i+1, qkey), stockId, flynumb, plateArea)
                plateDict = dict(zip(mapkeys, plate))
                platesMap[pmapKey] = plateDict
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #serialise dictionary
        picklepath = os.path.join(self.pickledir, 'platesMap_%s.pickle' %time)
        with open(picklepath, 'wb') as f:
            pickle.dump(platesMap, f, protocol = 2) 
        return
        
        
    def loadPlatesMap(self, batch = 'all'):
        import cPickle as pickle
        #Load areasDict
        filenames = [item for item in os.listdir(self.pickledir) if item.startswith('platesMap')]
        filenames.sort()
        picklefile = filenames[-1]
        picklePath = os.path.join(self.pickledir, picklefile) 
        with open(picklePath, 'rb') as f:
                platesMap = pickle.load(f)
        if isinstance(batch, (list, tuple, np.ndarray)):
            keyset = [[key for key in platesMap.keys() if key.split('_')[0]==str(batchnum)] for batchnum in batch]
            keyset = [key for batch in keyset for key in batch]#unpack list of lists
            platesMap = dict([(key, platesMap[key]) for key in keyset])     
        elif isinstance(batch, int):
            keyset = [key for key in platesMap.keys() if key.split('_')[0]==str(batch)] 
            platesMap = dict([(key, platesMap[key]) for key in keyset])
        return platesMap
    
            
    def dsetKeysMapper(self):
        '''It derives a batchlist from platesDir and returns a serialized dictionary of dictionaries object(dsetKeysMap.pickle). 
        Each dictionary contains deposit keys clustered according to deposit type and gender, batch, stock and plateIDs 
        {dtype&gender: {batch: {stock: {plateID: {depositKeys}}}}}.'''
        import cPickle as pickle
        from datetime import datetime
        #Define batchlist
        numbatches = len([item for item in os.listdir(self.platesdir) if item.startswith('QAFF')])
        batchlist = [i+1 for i in xrange(numbatches)]
        #batch: map keys 
        dsetKeys = [] 
        for batch in batchlist:    
            #Load dictionaries
            platesMap = self.loadPlatesMap(batch = batch)
            #Fetch and plate-cluster deposit keys
            print('\nClustering batch keys: QAFF%i' %batch)
            dkeysCluster = DataOperations().keysCluster(batch)
            #Unpack dkeys_plateClusters
            [rodkeysF, rodkeysM], [nrodkeysF, nrodkeysM] = rodkeys, nrodkeys = dkeysCluster
            dkeys_plateClusters = [rodkeysF, nrodkeysF, nrodkeysM]
            #Filter out empty plate clusters
            dkeys_plateClusters = [[plate for plate in depositType if len(plate)>0] for depositType in dkeys_plateClusters]
            #Stock-cluster dkeys_plateClusters and build a dictionary object from clusters
            stocksets = [list(set([platesMap['%s_%s' %(batch,plate[0].split('_')[1])]['stockId'][:-2] for plate in dType])) for dType in dkeys_plateClusters]
            stocknames = [['JS%i' %int(stock) if stock.isdigit() else stock for stock in sublist] for sublist in stocksets]    
            dkeys_stockClusters = [[(stocknames[i][j],[plate for plate in dkeys_plateClusters[i] if platesMap['%s_%s' %(batch, plate[0].split('_')[1])]['stockId'][:-2] == stock]) for j, stock in enumerate(sublist)] for i, sublist in enumerate(stocksets)]
            dkeys_stockClusters = [dict(depositType) for depositType in dkeys_stockClusters]
            dsetKeys.append(dkeys_stockClusters)
        #Zip batches and build a dictionary object    
        dsetKeys = zip(*dsetKeys)
        dsetKeys = [[(i+1, sublist) for i, sublist in enumerate(depositType)] for depositType in dsetKeys]
        dsetKeys = [dict(depositType) for depositType in dsetKeys]
        rodkeys, nrodkeysF, nrodkeysM = dsetKeys
        #Build a dictionary for depositTypes
        dsetKeysMap = {'rod': rodkeys, 'nrodF': nrodkeysF, 'nrodM': nrodkeysM}
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #Serialize dsetKeysMap
        path = os.path.join(self.pickledir, 'dsetKeysMap_%s.pickle' %time)
        with open(path, 'wb') as f:
            pickle.dump(dsetKeysMap, f, protocol = 2)
        return
    
    def loadDatasetKeysMap(self):
        import cPickle as pickle
        filenamelist = [filename for filename in os.listdir(self.pickledir) if filename.startswith('dsetKeysMap')]; filenamelist.sort()
        latestfile = filenamelist[-1]
        picklepath = os.path.join(self.pickledir, latestfile)
        with open(picklepath, 'rb') as f:
            dsetKeysmap = pickle.load(f)
        return dsetKeysmap
        
    def buildRODColourFilter(self):
        from datetime import datetime
        import cPickle as pickle
        '''It returns a dictionary object. Keys: batchnumber; values: (s, l) tuple for that particular batch.
        s and l values are calculated after '''
        rodColourFilter = {} 
        for batch in self.batchnumbers:
            try:
                print('Tuning ROD filter for batch%s' %batch)
                sl_tuple = RODFiltering().optimizeRODcFilter(batch, step = 10, maxval = [1.0, 0.6])
                rodColourFilter[batch] = sl_tuple[0]
            except Exception:
                print('Filtering parameters for batch%s were not computed.' %batch)
                continue
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #serialise dictionary
        filepath = os.path .join(self.pickledir, 'rodColourFilter_%s.pickle' %time)
        with open(filepath, 'w') as f:
            pickle.dump(rodColourFilter, f)
        return 
    
    def loadRODColourFilter(self):
        import cPickle as pickle
        filenamelist = [filename for filename in os.listdir(self.pickledir) if filename.startswith('rodColour')]; filenamelist.sort()
        latestfile = filenamelist[-1]
        picklepath = os.path.join(self.pickledir, latestfile)
        with open(picklepath, 'r') as f:
            rodColourFilter = pickle.load(f) 
        return rodColourFilter
    
    
    def buildRODSizeFilter(self, batchlim = [1,10]):
        import cPickle as pickle
        from itertools import chain
        from datetime import datetime
        from scipy import stats
        from astroML.stats import fit_bivariate_normal
        print('Fetching ROD population: QAFF%s to QAFF%s' %(batchlim[0], batchlim[1]))
        rodsize_dset = RODFiltering().fecthRODPopulation(batchlim = batchlim)
        #unpack data
        rodsizeF, rodsizeM = zip(*rodsize_dset)
        rodsizeF =zip(*rodsizeF)
        rodsizeM =zip(*rodsizeM)
        areaF, perimeterF = [np.asarray(list(chain(*sublist))) for sublist in rodsizeF]#reformat as numpy array
        #box cox transfomation
        x_cox, xlambda_cox = stats.boxcox(areaF)
        y_cox, ylambda_cox = stats.boxcox(perimeterF)
        print('Boxcox transformation: xlambda = %s, ylambda = %s' %(xlambda_cox, ylambda_cox))
        # compute non-robust and robust statistics
        (mu_nr, sigma1_nr,sigma2_nr, alpha_nr) = fit_bivariate_normal(x_cox, y_cox, robust=False)
        (mu_r, sigma1_r,sigma2_r, alpha_r) = fit_bivariate_normal(x_cox, y_cox, robust=True)
        print('Bivariate normal fit estimate: (%s, %s, %s, %s)' %(mu_nr, sigma1_nr,sigma2_nr, alpha_nr))
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #build dictionary object and serialise it
        sizeFilter = {'rob': (mu_r, sigma1_r,sigma2_r, alpha_r), 'nrob': (mu_nr, sigma1_nr,sigma2_nr, alpha_nr), 'boxcox': (xlambda_cox, ylambda_cox)}
        filepath = os.path.join(self.pickledir, 'rodSizeFilter_%s.pickle' %time)
        with open(filepath , 'w') as f:
            pickle.dump(sizeFilter, f)
        return
                
    def loadRODSizeFilter(self):
        import cPickle as pickle
        filenamelist = [filename for filename in os.listdir(self.pickledir) if filename.startswith('rodSize')]; filenamelist.sort()
        latestfile = filenamelist[-1]
        filepath = os.path.join(QAFFObjects().pickledir, latestfile)
        with open(filepath, 'r') as f:
            sizeFilter = pickle.load(f)
        return sizeFilter 
               
    def buildDatasetMetricsObj(self):
        from collections import OrderedDict
        import cPickle as pickle
        from datetime import datetime
        #fetch data objects
        workdset = QAFFObjects().loadWorkDataset()
        stockmetricsMap = DataOperations().loadStockMetricsMap()
        stockdatakeys = [mkeys, skeys] = DataOperations().fetchStockdataKeys()
        #build dataset metrics dictionary
        datasetMetrics = []
        metrics_mismatches = []
        for uId in workdset:
            try:
                print('Analysing %s' %uId)
                #fetch stockdata
                filename = stockmetricsMap[uId]
                filepath = os.path.join(QAFFObjects().stocksdir, filename)
                with open(filepath, 'r') as f:
                    stockdata = pickle.load(f)
                #test whether stockdata is empty 
                assert len(stockdata.values())>0, '%s: stockdata dictionary is empty.' %uId
                #unpack stockdata
                batchlist = stockdata.keys()
                stockdata = [stockdata[batch] for batch in batchlist]
                stockdata_means = [batch['means'] for batch in stockdata]
                stockdata_seq = [batch['seq'] for batch in stockdata]
                #reformat plate keys and add uId
                pkeys = [batch['seq']['pkeys'] for batch in stockdata]
                m_batchlists, s_batchlists = [[[int(batch[0][0].split('_')[0]) for batch in pkeys] for i in xrange(len(keylist))] for keylist in stockdatakeys]
                #zip keys and mean_metrics
                stockdata_means = [[batch[key] for key in mkeys] for batch in stockdata_means]
                stockdata_means = zip(*stockdata_means)
                stockdata_means = [zip(*metric) if isinstance(metric[0], (list, tuple)) else [metric] for metric in stockdata_means]
                stockdata_means = zip([uId]*len(mkeys), zip(m_batchlists, stockdata_means))
                #zip keys and seq_metrics
                stockdata_seq = [[batch[key] for key in skeys[1:]] for batch in stockdata_seq]
                stockdata_seq = [zip(*stockdata_seq)][0]
                stockdata_seq = [zip(*metric) if isinstance(metric[0], (list, tuple)) else [metric] for metric in stockdata_seq]
                stockdata_seq = zip([uId]*len(skeys), zip(s_batchlists, stockdata_seq))
                #stockdata
                stockdata = [stockdata_means, stockdata_seq]
                datasetMetrics.append(stockdata)
            except KeyError:
                print('%s has yet to be analysed' %uId)
                metrics_mismatches.append(uId)
                continue
            except AssertionError as ae:
                print(ae)
                continue
        #zip mkeys with clustered metrics 
        datasetMetrics = zip(*datasetMetrics)
        datasetMetrics = [zip(*datalist) for datalist in datasetMetrics]
        datasetMetrics = [[OrderedDict(metric) for metric in datalist]for datalist in datasetMetrics]   
        stockdatakeys = [mkeys, skeys[1:]]
        datasetMetrics = [dict(zip(stockdatakeys[i], datalist)) for i, datalist in enumerate(datasetMetrics)]
        datasetMetrics_keys = ['means', 'seq']
        datasetMetrics = dict([(datasetMetrics_keys[i], sublist) for i, sublist in enumerate(datasetMetrics)])
        #fetch current time
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #serialise datsetMetrics
        picklepath = os.path.join(self.pickledir, 'datasetMetrics_%s.pickle' %time)
        with open(picklepath, 'wb') as f:
            pickle.dump(datasetMetrics, f, protocol = 2)
        #save mismatches
        if len(metrics_mismatches)>0:
            lines = ('\n').join(metrics_mismatches)
            mismatchdir = os.path.join(self.workdir, 'Mismatches')
            filepath = os.path.join(mismatchdir, 'metrics_mismatches_%s.txt' %time)
            with open(filepath, 'w') as f:
                f.writelines(lines)
        return

    def loadDatasetMetrics(self):
        import cPickle as pickle
        filenamelist = [filename for filename in os.listdir(self.pickledir) if filename.startswith('datasetMetrics')]; filenamelist.sort()
        latestfile = filenamelist[-1]
        picklepath = os.path.join(self.pickledir, latestfile)
        with open(picklepath, 'rb') as f:
            datasetMetrics = pickle.load(f)
        return datasetMetrics
    
    def buildBatchMetricsObj(self):
        from itertools import chain
        from datetime import datetime
        import cPickle as pickle
        #fetch data objects
        datasetMetrics = self.loadDatasetMetrics()
        #fetch dataset metrics 
        metrickeys = datasetMetrics['seq'].keys()
        metricsvalues = [datasetMetrics['seq'][key].values() for key in metrickeys]
        #define data variables
        dsetbatches = [i for i in xrange(22)]
        dtypekeys = ['rod', 'females', 'males']
        #reformat datasetMetrics: cluster data on batches
        batchMetrics_list = [[[[] for i in dsetbatches] for j in dtypekeys] for metric in metrickeys]
        for z, metric in enumerate(metricsvalues):
            for stock in metric:
                batches = stock[0]
                for i, batch in enumerate(batches):
                    if metrickeys[z] == 'rodFraction':
                        batchMetrics_list[z][0][batch-1].append(stock[1][0][i])
                    else:
                        [batchMetrics_list[z][j][batch-1].append(stock[1][j][i]) for j, key in enumerate(dtypekeys)]
        #build batch metrics dictionary
        batchMetrics_list = [[[list(chain(*batch)) for batch in dtype] for dtype in metric] for metric in batchMetrics_list]#unpack batches
        batchMetrics_list = [[[list(chain(*batch)) if len(batch)==0 or isinstance(batch[0], list) else batch for batch in dtype] for dtype in metric] for metric in batchMetrics_list]#unpack batches
        batchMetrics_list = [dict(zip(dtypekeys, metric)) for metric in batchMetrics_list]
        batchMetrics = {}
        for i, key in enumerate(metrickeys): 
            batchMetrics[key] = batchMetrics_list[i]
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')   
        #serialise dictionary
        picklepath = os.path.join(Dashboard().pickledir, 'batchMetrics_%s.pickle' %time)
        with open(picklepath, 'wb') as f:
            pickle.dump(batchMetrics, f, protocol = 2)
        return
    
    def loadBatchMetrics(self):
        import cPickle as pickle
        '''Loads the latest version of batchMetrics.'''
        #Load sqlkeysDict
        filenames = [item for item in os.listdir(self.pickledir) if item.startswith('batchMetrics')]
        filenames.sort()
        picklefile = filenames[-1]  
        picklePath = os.path.join(self.pickledir, picklefile)
        with open(picklePath, 'rb') as f:
                batchMetrics = pickle.load(f)      
        return batchMetrics
    
    def buildFDRforRODsObject(self):
        import cPickle as pickle
        from datetime import datetime
        rodFDR = RODFiltering().calculateFDRforRODs()
        #fetch current time
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        picklepath = os.path.join(self.pickledir, 'rodFDR_%s.pickle' %time)
        with open(picklepath, 'w') as f:
            pickle.dump(rodFDR, f) 
        return
        
    def loadFDRforRODs(self):
        import cPickle as pickle
        filenamelist = [filename for filename in os.listdir(self.pickledir) if filename.startswith('rodFDR')]; filenamelist.sort()
        latestfile = filenamelist[-1]
        filepath = os.path.join(self.pickledir, latestfile)
        with open(filepath, 'r') as f:
            rodFDR = pickle.load(f)
        return rodFDR
        
    def buildTypeIRODkeysDict(self):
        import cPickle as pickle
        from datetime import datetime
        typeIRODkeys = RODFiltering().fetchTypeIRODkeys()
        #fetch current time
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #serialise dictionary
        picklepath = os.path.join(self.pickledir, 'typeIRODkeys_%s.pickle' %time)
        with open(picklepath, 'wb') as f:
            pickle.dump(typeIRODkeys, f, protocol = 2)
        return
    
    def loadTypeIRODkeys(self):
        import cPickle as pickle
        filenamelist = [filename for filename in os.listdir(self.pickledir) if filename.startswith('typeIRODkeys')]; filenamelist.sort()
        latestfile = filenamelist[-1]
        filepath = os.path.join(self.pickledir, latestfile)
        with open(filepath, 'rb') as f:
            typeIRODkeys = pickle.load(f)
        return typeIRODkeys

        

class DataOperations(QAFFObjects):
    
    def __init__(self):
        QAFFObjects.__init__(self)
        self.mismatchesdir = os.path.join(self.workdir, 'Mismatches')
        
    
    def getMembers(self):
        import inspect
        methods = inspect.getmembers(self, predicate = inspect.ismethod)
        print(methods)
        return  
            
    def fetchGridmapMismatches(self):
        from itertools import ifilter
        from datetime import datetime
        '''It fetches the gridmap sequence of qkeys and returns the list of qkeys that are not present in areasDict (areaMismatches)
        and the list of qkeys that are not present in sqlkeysDict (voidMismatches).'''
        #Load gridmapDict
        gridmapDict = self.loadGridmap()
        #Fetch qkeys from gridmapDict
        batchnames = gridmapDict.keys()
        batchnames = sorted(batchnames, key = lambda x:int(x[4:]))#sort batchnames 
        gridmap_qkeys = [[qkey for (qkey, stockId, flynumb) in gridmapDict[name]] for name in batchnames]
        #Load areas and sqlKeys dictionaries
        sqlKeysDict = self.loadSQLkeysDict()
        areasDict = self.loadAreasDict()
        #Find area and void mismatches
        areaMismatches = [['%i%s' %(i+1, qkey) for qkey in ifilter(lambda x: (self.fromQkeyToScan(x)[0] not in areasDict['QAFF%i' %(i+1)].keys()), batch)] for i, batch in enumerate(gridmap_qkeys)]
        voidMismatches = [['%i%s' %(i+1, qkey) for qkey in ifilter(lambda x: (self.fromQkeyToScan(x)[0] not in sqlKeysDict['QAFF%i' %(i+1)].keys()), batch)] for i, batch in enumerate(gridmap_qkeys)]
        #Unpack mismatches
        areaMismatches = [qkey for batch in areaMismatches for qkey in batch]
        voidMismatches = [qkey for batch in voidMismatches for qkey in batch]
        mismatches = [areaMismatches, voidMismatches]
        #fetch current time 
        time = time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #save mismatches to file
        lines = ['\n'.join(sublist) for sublist in mismatches]
        filenames = ['plates_area_mismatches', 'plates_void_mismatches']; filenames = ['%s_%s.txt' %(filename, time) for filename in filenames]
        filepaths = [os.path.join(self.mismatchesdir, filename) for filename in filenames]
        for i, filepath in enumerate(filepaths):
            with open(filepath, 'w') as f:
                f.writelines(lines[i])
        
        return areaMismatches, voidMismatches


    def printPlatesMap(self, batch = 'all'):
        from datetime import datetime
        #load dictionary
        platesMap = self.loadPlatesMap(batch = batch)
        platesMap_keys = ['qkey', 'stockId', 'flycount', 'area']
        #Fetch rows from platesMap
        rows = [(sqlkey.split('_')[0], sqlkey.split('_')[1], [platesMap_val[key] for key in platesMap_keys]) for (sqlkey, platesMap_val) in platesMap.items()]
        rows = sorted(rows, key = lambda x:int(x[0])) # sort rows according to catch number
        #Assemble lines from rows
        lines = ['%s\t%s\t%s\n' %(row[0], row[1], '\t'.join(map(str,row[2]))) for row in rows]
        lines.insert(0, 'Batch\tSQLkey\tQkey\tStockId\tFly count\tPlate area\n')
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        #Write lines to file
        textdir = os.path.join(self.workdir, 'TextFiles')
        path = os.path.join(textdir, 'platesMap_%s.txt' %time)
        with open(path, 'w') as f:
            f.writelines(lines)
        return
    
    
    def purgePickleDir(self, fnum = 2):
        '''It purges pickledir of older file versions; fnum (int) defines the maximum number of file versions to keep.
        Pickle files purged are: gridMap, platesMap, areasDict, sqlkeysDict and dsetKeysMap files.'''
        #Define filenames
        filenames = ['gridMap', 'platesMap', 'areasDict', 'sqlkeysDict', 'dsetKeysMap', 'datasetMetrics', 'rodColour', 'rodSize', 'batchMetrics', 'typeIRODkeys', 'rodFDR']
        #Fetch and sort filelist
        filelist = [[f for f in os.listdir(self.pickledir) if f.startswith(name)] for name in filenames]
        [sublist.sort() for sublist in filelist] #sort filelist
        #Purge filelist
        filelist_topurge = [sublist[:-fnum] for sublist in filelist]
        #Purge pickleDir
        [[os.remove(os.path.join(self.pickledir, filename)) for filename in sublist] for sublist in filelist_topurge]

        return
        
    def purgeMismatchesDir(self, fnum = 2):
        '''It purges mismatchesdir of older file versions; fnum (int) defines the maximum number of file versions to keep.
        Pickle files purged are: bug', iQAFF, malefemale, metrics and void mismatch files.'''
        #Define filenames
        filenames = ['bug', 'iQAFF', 'malefemale', 'metrics', 'plates_void', 'plates_area']
        #Fetch and sort filelist
        filelist = [[f for f in os.listdir(self.mismatchesdir) if f.startswith(name)] for name in filenames]
        [sublist.sort() for sublist in filelist] #sort filelist
        #Purge filelist
        filelist_topurge = [sublist[:-fnum] for sublist in filelist]
        #Purge mismatchesDir
        [[os.remove(os.path.join(self.mismatchesdir, filename)) for filename in sublist] for sublist in filelist_topurge]

        return
    
    
    def fromSqlKeyToScan(self, sqlkeys, batch):
        '''It takes a sqlKey or a sequence of sqlKeys (tuple, list) as input and returns the corresponding scanID(s). Batch (int) 
        indicates the batchnumber. '''
        #load platesMap
        platesMap = self.loadPlatesMap(batch = batch)
        #fetch scanID
        if isinstance(sqlkeys, int):
                sqlkeys = [sqlkeys]
        elif isinstance(sqlkeys, (list, tuple, np.ndarray)):
            try:
                qkeys = [platesMap['%i_%s' %(batch, key)]['qkey'] for key in sqlkeys]
                scanids = [self.fromQkeyToScan(key) for key in qkeys]
                scanids = [item for sublist in scanids for item in sublist]
            except KeyError:
                print('\nKeyError: sqlkeys.')
                return    
        else:
            raise TypeError('\n sqlkeys type is not recognised.')
        return scanids    


    def fromQkeyToSQLkey(self, qkeys, batch):
        '''It takes a qKey or a sequence of qKeys (tuple, list) as input and returns the corresponding sqlKey(s). Batch (int) 
        indicates the batchnumber. '''
        #load plates map
        platesMap = self.loadPlatesMap(batch = batch)
        #reformat datatype
        if isinstance(qkeys, str):
            qkeys = [qkeys]
        #fetch sqlkeys
        sqlkeys = [[key for (key, values) in platesMap.items() if values['qkey'] == qkey] for qkey in qkeys]
        sqlkeys = [key for batch in sqlkeys for key in batch] 
        if len(sqlkeys)==1:
            sqlkeys = sqlkeys[0]    
        return sqlkeys 
     
    
    def fromScanToQkey(self, filenamelist, batch = None):
        '''It takes a scan filename or a sequence of filenames (tuple, list) and returns the corresponding Qkey(s). If
        batch = batchnumber (int, str) it appends batchnumber as a prefix to the corresponding Qkey(s).'''
        if isinstance(filenamelist, (tuple, list, np.ndarray)):
            nameparts = [filename.split('_') for filename in filenamelist]
            scanid, gridxy = zip(*nameparts)
            scanid = ['%s%i' %(item[0], int(item[4:])) for item in scanid]
            cols, rows = zip(*[(item[0], item[1:]) for item in gridxy])
            Qkeys = [''.join(sublist) for sublist in zip(scanid, cols, rows)]                
        elif isinstance(filenamelist, str):
            scanid, gridxy = filenamelist.split('_')
            scanid = '%s%i' %(scanid[0], int(scanid[4:]))
            cols, rows = gridxy[0], gridxy[1:]
            Qkeys = [''.join([scanid,cols, rows])]        
        if isinstance(batch, (str, int)):
                Qkeys = ['%s%s' %(batch, key) for key in Qkeys]             
        else:
            raise TypeError('\n Batch must be str or int type. ')      
        return Qkeys  


    def fromQkeyToScan(self, keylist):
        '''It takes a Qkey or a sequence of Qkeys (tuple, list) and returns the corresponding scanID(s).''' 
        #reformat keyword datatype    
        if isinstance(keylist, str):
                keylist = [keylist]
        #fetch scanIDs
        idx = [[i for i, val in enumerate(key) if not val.isdigit()] for key in keylist]
        keyparts = [(key[:idx[i][0]], int(key[idx[i][0]+1:idx[i][1]]), key[idx[i][1]], key[idx[i][1]+1:]) for i, key in enumerate(keylist)]
        scanids = ['Scan%.3i_%s%s' %(part[1], part[2], part[3]) for part in keyparts]    
        return scanids


    def depositsDict(self, batch):
        '''It takes a batch number (int) as input and returns a dictionary object containing all the metrics 
        for each deposit in that batch. Keys: 'plateID_depositID' (string); values: Xposition, Yposition, Area, Perimeter, Circularity,
        ROD, IOD, MeanB, MeanG, MeanR, MeanH, MeanL, MeanS. '''
        #test arg datatype
        if isinstance(batch, int):
            rawfile = 'QAFF%i_deposits.txt' %batch
        else:
            raise TypeError('\nBatch var must be type int. ')
        #create dictionary    
        rawpath = os.path.join(self.depositsdir, rawfile) 
        depositsDict = dictFromFile(rawpath, [0,1], separator = '_', colasarr = True) 
        return depositsDict


    def platesDict(self, batch):
        '''It takes a batch number (int) as input and returns a dictionary object containing all the metrics 
        for each plate in that batch. Keys: sqlKey (int); values: plateName, imageFileName, groupName and
        plate metrics - meanArea, meanPerimeter, meanCircularity, totalIOD, meanH, meanL, meanS, etc. '''
        #test arg datatype
        if isinstance(batch, int):
            platefile = 'QAFF%i_plates.txt' %batch
        else:
            raise TypeError('\nBatch var must be type int. ')    
        #create dictionary    
        platepath = os.path.join(self.platesdir, platefile)
        platesDict = dictFromFile(platepath, 0, colasarr = True)
        return platesDict
            

    def keysCluster(self, batch):
        from File_Functions import listPartition
        from statsFunctions import pointInEllipse
        '''It takes a batch number as input and clusters the respective deposit keys according to deposit type, 
        gender and plateID. ROD keys are filtered on the basis of colour and size of RODs. Colour filtration uses the saturation and lightness values defined in lsDict for each batch.
        Size filtration uses the filter parameters defined in rodsizeFilter.'''
        #Load dictionaries
        platesMap = QAFFObjects().loadPlatesMap(batch = batch)
        depositsDict = DataOperations().depositsDict(batch)
        platesDict = DataOperations().platesDict(batch)
        lsDict = QAFFObjects().loadRODColourFilter()
        rodsizeFilter = QAFFObjects().loadRODSizeFilter()
        #Filter out headings from keyset
        keyset = [key for key in depositsDict.keys() if key not in Dashboard().dheads]   
        #Fetch ROD and non-ROD depositkeys
        rodkeys, nrodkeys = listPartition(lambda x: depositsDict[x]['ROD'] == 1, keyset)
        #Filter ROD keys using color metrics
        rodLS = [(key, depositsDict[key]['meanL'], depositsDict[key]['meanS']) for key in rodkeys] ##Fetch lightness and saturation values
        s, l  = lsDict[batch] ##Fetch filter values  
        rodkeys_filtered = [tupl[0] for tupl in rodLS if tupl[1] <= l and tupl[2] <= s]
        #Fetch RODs size data
        rodSize = [(depositsDict[key]['area'], depositsDict[key]['perimeter']) for key in rodkeys_filtered]
        area, perimeter = zip(*rodSize)#unpack data
        #apply box cox transformation to data
        xlambda, ylambda = rodsizeFilter['boxcox']
        xcox = (area**xlambda - 1) / xlambda 
        ycox = (perimeter**ylambda - 1) / ylambda
        #fetch normal bivariate fits
        (mu_nr, sigma1_nr,sigma2_nr, alpha_nr) = rodsizeFilter['nrob']
        #test whether points are inside elliptic surface
        spread = RODFiltering().sfilterSpreads[batch]
        sizefiltered = [pointInEllipse(mu_nr[0], mu_nr[1], valx, ycox[i], sigma1_nr*spread, sigma2_nr*spread, alpha_nr) for i, valx in enumerate(xcox)]
        sizefiltered = np.asarray(sizefiltered); rodkeys_filtered = np.asarray(rodkeys_filtered)
        #filter out putative ROD false positives by size
        rodkeys_2xfiltered = rodkeys_filtered[sizefiltered]
        depositKeys = [rodkeys_2xfiltered, nrodkeys]
        #Cluster deposits for each plate
        batchsize = len([key for key in platesDict.keys() if isinstance(key, (np.int64, int, float))])
        depositkeys_clustered = [[[] for i in xrange(batchsize)], [[] for i in xrange(batchsize)]]
        [[depositkeys_clustered[i][int(key.split('_')[1])-1].append(key) for key in sublist] for i, sublist in enumerate(depositKeys)]
        rodkeys_clustered, nrodkeys_clustered = depositkeys_clustered    
        #Cluster plate keys on the basis of fly gender
        batchkeys = platesMap.keys()          
        batchkeys_partition = listPartition(lambda x:platesMap[x]['stockId'][-2] == 'F', batchkeys)
        depositkeys_clustered = [[[dtype[int(batchkey.split('_')[1])-1] for batchkey in sublist] for sublist in batchkeys_partition] for dtype in depositkeys_clustered]
        return depositkeys_clustered

        
    def fetchStockdataKeys(self):
        mkeys = ['meanCount', 'errCount', 'mean_rodFraction', 'err_rodFraction', 'size', 'errSize', 'hue', 'errH', 'lightness', 'errL', 'saturation', 'errS']
        skeys = ['pkeys','pooCount', 'normCount', 'rodFraction', 'seqA', 'seqH', 'seqL', 'seqS']
        stockdataKeys = [mkeys, skeys]
        return stockdataKeys
        
    def calculateStockMetrics(self, stockinput):
        import cPickle as pickle
        from tempfile import NamedTemporaryFile
        from itertools import ifilter
        from scipy.stats import sem
        '''It takes a stockID as input and returns a serialized dictionary object that contains the metrics 
        for that stock. Updates a filemap dictionary containing the filepaths of all stockmetrics dictionaries already calculated; entries are time-stamped.'''
        #Define batchlist
        numbatches = len([item for item in os.listdir(self.platesdir) if item.startswith('QAFF')])
        batchlist = [i+1 for i in xrange(numbatches)]
        #Load dsetKeyMap dictionary
        dsetKeysMap = self.loadDatasetKeysMap()
        #Define key combinations  
        dtypeKeys = ['rod', 'nrodF', 'nrodM']
        keytuples = [[[(dtype, batch, stock) for stock in ifilter(lambda x: x == stockinput, dsetKeysMap[dtype][batch].keys())] for batch in batchlist] for dtype in dtypeKeys]
        keytuples = [[item[0] for item in sublist if len(item) > 0] for sublist in keytuples]
        if len(zip(*keytuples)) == 0:
            raise  TypeError('dsetKeysMap[%s] returns an empty list.' %stockinput)
        #Test whether males and females were assayed simultaneously
        if [item[1] for item in keytuples[0]] == [item[1] for item in keytuples[2]]:
            batches = [[item[1] for item in keytuples[0]]]
        else:
            batches = [[item[1] for item in keytuples[0]], [item[1] for item in keytuples[2]]]

        paramDict = {} 
        #Calculate poo metrics for each stock
        if len(batches) == 1: 
            for batch in batches[0]:
                #Fetch keymaps
                keyClusters = [dsetKeysMap['rod'][batch][stockinput], dsetKeysMap['nrodF'][batch][stockinput], dsetKeysMap['nrodM'][batch][stockinput]]
                #Load dictionaries
                platesMap = self.loadPlatesMap(batch = batch)
                depDict = self.depositsDict(batch)
                #Calculate plates deposit count: RODs and non-RODs
                plateArea_max = 4392897
                pkeys = [[int(plate[0].split('_')[1]) for plate in dtypeCluster] for dtypeCluster in keyClusters]
                batchPkeys = [['%s_%s' %(batch, pkey) for pkey in sublist] for sublist in pkeys]
                flyCount = [[platesMap[pkey]['flycount'] for pkey in sublist] for sublist in batchPkeys]
                plateArea = [[platesMap[pkey]['plate area'] for pkey in sublist] for sublist in batchPkeys]
                #metrics: count
                pooCount = [[len(plate) for plate in sublist] for sublist in keyClusters]
                rodFraction = np.asarray(pooCount[0], dtype = 'f')/sum(pooCount[0] + pooCount[1])
                normCount = [[(count/float(flyCount[i][j])/(plateArea[i][j]/float(plateArea_max))) for j, count in enumerate(sublist)] for i, sublist in enumerate(pooCount)]
                #Determine stock meanCount: RODs and non-RODs 
                meanCount = [np.mean(sublist) for sublist in normCount]
                errCount = [sem(sublist) for sublist in normCount]
                mean_rodFraction = np.mean(rodFraction)
                err_rodFraction = sem(rodFraction)
                #Fetch plate area, hue and lightness values sequences
                seqA = [[[depDict[key]['area'] for key in plate] for plate in sublist] for sublist in keyClusters]
                seqH = [[[depDict[key]['meanH'] for key in plate] for plate in sublist] for sublist in keyClusters]
                seqL = [[[depDict[key]['meanL'] for key in plate] for plate in sublist] for sublist in keyClusters] 
                seqS = [[[depDict[key]['meanS'] for key in plate] for plate in sublist] for sublist in keyClusters]
                #Calculate means and errors
                size = [[np.mean(plate) for plate in sublist] for sublist in seqA]
                errSize = [[sem(plate) for plate in sublist] for sublist in seqA]
                hue = [[np.mean(plate) for plate in sublist] for sublist in seqH]
                errH = [[sem(plate) for plate in sublist] for sublist in seqH]
                lightness = [[np.mean(plate) for plate in sublist] for sublist in seqL]
                errL = [[sem(plate) for plate in sublist] for sublist in seqL]
                saturation = [[np.mean(plate) for plate in sublist] for sublist in seqS]
                errS = [[sem(plate) for plate in sublist] for sublist in seqS]
                #Generate poo metrics subdictionaries and append them to paramDict
                stockdataKeys = self.fetchStockdataKeys()
                mkeys, skeys = stockdataKeys
                meansDict = dict(zip(mkeys, [meanCount, errCount, mean_rodFraction, err_rodFraction, size, errSize, hue, errH, lightness, errL, saturation, errS]))
                seqDict = dict(zip(skeys, [batchPkeys, pooCount, normCount, rodFraction, seqA, seqH, seqL, seqS]))
                paramDict[batch] = dict([('means', meansDict), ('seq', seqDict)])
                
        elif len(batches) == 2:
            raise ValueError('%s: males and females were not assayed simultaneously.' %stockinput)
        #test whether Stocks directory exists
        stockdir = os.path.join(self.pickledir, 'Stocks')
        if not os.path.isdir(stockdir):
            os.mkdir(stockdir)
        #serialize stock metrics dictionary     
        with NamedTemporaryFile('w+t', prefix = ('%s_tmp_' %stockinput), dir = stockdir, delete = False) as f:
            pickle.dump(paramDict, f)
            picklefile = f.name   
        #Search for old stockinput tempfiles and delete them    
        old_tempfiles = [filename for filename in os.listdir(stockdir) if filename.split('_')[0] == stockinput and filename != os.path.split(picklefile)[1]]
        if len(old_tempfiles)>0:
            [os.remove(os.path.join(stockdir, filename)) for filename in old_tempfiles]
            #Update stockmetrics dictionary
            filelist = os.listdir(stockdir)
            tempDict = dict([(filename.split('_')[0], filename) for filename in filelist if filename != 'stockmetricsMap.pickle'])
            stockmetricsMap = os.path.join(stockdir, 'stockmetricsMap.pickle')
            with open(stockmetricsMap, 'w') as  f:
                pickle.dump(tempDict, f)        
        return 
        
    
    def loadStockMetricsMap(self):
        import cPickle as pickle
        #load dictionary
        picklepath = os.path.join(self.stocksdir, 'stockmetricsMap.pickle')
        with open(picklepath, 'r') as f:
            stockMetricsMap = pickle.load(f)
        return stockMetricsMap
         
         
    def stockInWorkDataset(self, stockinput):
        '''It tests whether a stock (uId) belongs to the working dataset. It returns an AssertionError if it does not belong.'''
         #test whether stockinput is in the work dataset
        stockinput = stockinput.upper()
        try:
            ctrlDict = self.controlsDict()
            stockinput = ctrlDict[stockinput]
        except KeyError:
            workdset = self.loadWorkDataset()
            assert stockinput in workdset, '%s is not part of the QAFF dataset.' %stockinput
        return stockinput

        
    def stockDataFetcher(self, stockinput):
        import cPickle as pickle
        '''It fetches the stockdata for a given stock (stockinput) from stockMetricsMap. If the stock metrics has not been calculated yet, it calculates it, 
        updates stockMetricsMap and returns it.'''
        #test whether stockinput is in the work dataset
        stockinput = self.stockInWorkDataset(stockinput)    
        #fetch stockdata
        while stockinput:
            stockmetricsMap = self.loadStockMetricsMap()#load file metrics map
            try:
                filename = stockmetricsMap[stockinput]#fetch stockdata
                filepath = os.path.join(self.stocksdir, filename)
                try:
                    with open(filepath, 'r') as f:
                        stockdata = pickle.load(f)
                        break
                except IOError:
                    self.calculateStockMetrics(stockinput)#calculate stockdata    
            except KeyError:
                self.calculateStockMetrics(stockinput)#calculate stockdata
        return stockdata
        
        
    def calculateMetricsForDataset(self):
        '''It calculates the metrics for each stock in the working dataset and controls. It captures stocks that are part of the working dataset 
        but have yet to be processed (workdatset_mismatches), stocks where males and females were assayed separately instead of simultaneously (mf_mismatches)
        and stocks that raise unspecific mismatches (bug_mismatches). It saves these lists separately as time-stamped text files in the Mismatches directory.'''
        from Unknome_Functions import Unknome
        from datetime import datetime
        #load dataset IDs
        uIDs = Unknome().uIDs
        uIDs = uIDs + self.ctrlnames
        workdataset_mismatches = []
        mf_mismatches = []
        bug_mismatches = []
        for uId in uIDs:
            try:
                self.stockInWorkDataset(uId)#test whether stock is in workdataset
                print('Analysing %s.' %uId)
                self.calculateStockMetrics(uId)
            except AssertionError:
                if uId in Unknome().fetchViables():
                    print('%s is a viable but not in the working dataset.' %uId)
                    workdataset_mismatches.append(uId)
                elif uId in Unknome().fetchLethals():
                    print('%s is a lethal.' %uId)
                elif uId not in self.ctrlnames:
                    print('IMPORTANT: %s is part of the working dataset but it is not a control or, a lethal or a viable.' %uId)
                    bug_mismatches.append(uId)           
            except ValueError as ve:
                print(ve)
                mf_mismatches.append(uId)
            except TypeError as kme:
                print(kme)
            continue
        #Fetch current time and date
        time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')    
        #save mismatches lists
        mismatchlists = [workdataset_mismatches, mf_mismatches, bug_mismatches]        
        lines = [('\n').join(lst) for lst in mismatchlists if len(lst)>0] 
        mismatchdir = os.path.join(self.workdir, 'Mismatches')
        filenames = ['wdataset_mismatches', 'malefemale_mismatches', 'bug_mismatches']
        filepaths = [os.path.join(mismatchdir, '%s_%s.txt' %(filename, time)) for i,filename in enumerate(filenames) if len(mismatchlists[i])>0]
        for i, path in enumerate(filepaths):
            with open(path, 'w') as f:
                f.writelines(lines[i])
        return
        


class PooDataVis(DataOperations):
    
    def __init__(self):
        DataOperations.__init__(self)
        self.dropimagedir = os.path.join(self.workdir, 'DropImage')
        if not os.path.exists(self.dropimagedir):
                os.makedirs(self.dropimagedir)
        
    def getMembers(self):
        import inspect
        methods = inspect.getmembers(self, predicate = inspect.ismethod)
        print(methods)
        return
        
    def plotBatchMeanAreas(self):
        import matplotlib.pyplot as plt
        #load areas object
        areasDict = self.loadAreasDict()
        batchnumbers = areasDict.keys()
        areas = areasDict.values()
        #calculate batch mean areas
        areas = [batch.values() for batch in areas] 
        areas = [np.mean(batch) for batch in areas]
        #plot areas
        ax = plt.subplot(111)
        xdataset = np.arange(len(batchnumbers))
        ax.plot(xdataset, areas)
        ax.set_xticks(xdataset)
        ax.set_xticklabels(batchnumbers, rotation = -90)
        ax.set_ylim([0, max(areas)+500000])
        plt.show()
        return
        
                
    def stockPlotter(self, stockinput, parameter, plotType):
        '''It takes a stock and plots the count, size, hue or lightness for each deposit type in each gender. Size, hue and lightness can
        be plotted for either the dataset, batch or plate.'''
        #Load stockdata
        paramDict = self.stockDataFetcher(stockinput)   
        #Define batches and build plate-keys dictionary
        batches = paramDict.keys(); numbatches = len(batches)
        pkeys = [paramDict[batch]['seq']['pkeys'] for batch in batches]
        pkeys = zip(*pkeys)   
        pkeys = [(batch, dict([('rod', pkeys[0][i]), ('females', pkeys[1][i]), ('males', pkeys[2][i])])) for i, batch in enumerate(batches)]
        pkeysDict = dict(pkeys)
        #test whether stock is in working dataset
        stockinput = self.stockInWorkDataset(stockinput)

        #Define keys and build datakeys dictionaries
        mkeys = ['meanCount', 'errCount', 'mean_rodFraction', 'err_rodFraction', 'size', 'errSize', 'hue', 'errH', 'lightness', 'errL', 'saturation', 'errS']
        mkeysDict = dict([(key, i) for i, key in enumerate(mkeys)])
        skeys = ['pkeys', 'pooCount', 'normCount', 'rodFraction', 'seqA', 'seqH', 'seqL', 'seqS']
        skeysDict = dict([(key, i) for i, key in enumerate(skeys)])
        
        #Fetch means and seq data from paramDict
        mData = [[paramDict[batch]['means'][key] for key in mkeys ] for batch in batches]
        sData = [[paramDict[batch]['seq'][key] for key in skeys ] for batch in batches]
        
        #Sort data: seqData = [(dpType[batch[plate]])]; meansData = [(dpType[batch])] 
        mData = zip(*mData)
        sData = zip(*sData)
        mData = [zip(*sublist) if isinstance(sublist[0], (list, tuple)) else sublist for sublist in mData] #[(dpType[batch])]; except mean_rodFraction = (batch)
        sData = [zip(*sublist) if isinstance(sublist[0], (list, tuple)) else sublist for sublist in sData ] #[(dpType[batch[plate]])]; except rodFraction = (batch[plate])
        
        platesMap = self.loadPlatesMap() #Load plates map
        args = [pkeysDict, mkeysDict, skeysDict, platesMap, sData, mData, batches, numbatches]
        
        if parameter == 'count':
            self.countPlotter(stockinput, args)
        else:
            self.hlsPlotter(stockinput, parameter, plotType, args) 
        return args
        
                  
    def countPlotter(self, stockinput, args):
        from matplotlib import pyplot as plt
        from mpldatacursor import datacursor
        import seaborn as sns
        '''It takes a stockID as input and returns a plot of the count for each deposit type in each gender. The count
        is plotted per batch and plate.'''
        
        pkeysDict, mkeysDict, skeysDict, platesMap, sData, mData, batches, numbatches = args
        
        #define figure
        plt.close('all')
        plt.figure('%s: deposit count' %stockinput)
        #Define plot position in the grid
        ax1 = plt.subplot2grid((1,3), (0,0), colspan=1)
        ax2 = plt.subplot2grid((1,3), (0,1), colspan=1)
        ax3 = plt.subplot2grid((1,3), (0,2), colspan=1)
        axes = [ax1, ax2, ax3]
        
        #Fetch ydata
        count = sData[skeysDict['normCount']]
        meanCount = mData[mkeysDict['meanCount']]
        errCount = mData[mkeysDict['errCount']]
        count = [[plate for batch in dtype for plate in batch] for dtype in count]#Flatten data
        #Unpack data
        rodCount, countF, countM = count
        meanrodCount, meanCountF, meanCountM = meanCount
        err_rodCount, errCountF, errCountM = errCount
        
        #Number of plates per batch
        platekeys = ['rod', 'females', 'males']
        platesPerBatch = [[len(pkeysDict[batch][key]) for batch in batches] for key in platekeys]
    
        #Define xdata
        xlabels = ['QAFF%s' %batch for batch in batches] 
        xdset_labels = [[[xlabels[i]]*val for i,val in enumerate(sublist)] for sublist in platesPerBatch]
        xdset_labels = [[label for batch in sublist for label in batch] for sublist in xdset_labels]#unpack
        
        #Define plots
        [sns.boxplot(x = xdset_labels[i], y = count[i], ax = ax) for i, ax in enumerate(axes)]
        [sns.swarmplot(x = xdset_labels[i], y = count[i], s = 4, color = 'b', ax = ax) for i, ax in enumerate(axes)]
        
        #fetch plate IDs
        plateIDs = [[[plate for plate in pkeysDict[batch][key]] for i, batch in enumerate(batches)] for key in platekeys]
        #Convert plateIDs to scanids
        scanids = [[[platesMap[label]['qkey'] for label in batch] for batch in dtype] for dtype in plateIDs]
        
        #label datapoints interactively
        datapoint_sets = [ax.collections[:] for ax in axes]
        for i, dataset in enumerate(datapoint_sets):
            for j, batch in enumerate(dataset):
                datacursor(batch, hover=True, point_labels = scanids[i][j], fontsize = 10, bbox= None, xytext=(0, 25), formatter=lambda **kwargs: kwargs['point_label'][0])
        
        #Define xticks labels
        [ax.set_xticklabels(xlabels, rotation = 60, size = 10) for ax in axes]
        #Set axes title and axis labels
        dtype_labels = ['ROD', 'nROD females', 'nROD males']
        [ax.set_title('%s: %s' %(stockinput, dtype_labels[i]), fontsize = 12) for i, ax in enumerate(axes)]
        ax1.set_ylabel('Number of deposits per fly per plate', fontsize = 13)
        #set axis limits
        ax1.set_ylim([0, max(rodCount)+2])
        
        #save plot
        filepath = os.path.join(self.dropimagedir, '%s_count.png' %stockinput)
        plt.savefig(filepath, bbox_inches='tight')
        
        plt.show()
        return
    
        
    def hlsPlotter(self, stockinput, parameter, plotType, args, numbins = 30, minval = 0):
        from Plot_Functions import hsvGenerator
        from colorsys import hsv_to_rgb
        from matplotlib import pyplot as plt
        from matplotlib.collections import PolyCollection
        from matplotlib.colors import colorConverter
        from mpl_toolkits.mplot3d import Axes3D
        from itertools import chain
        from math import sqrt
        from scipy.stats import sem
        '''It takes a stockID as input and returns a plot of either the deposit size, hue or lightness for each deposit type in each gender. It plots
        each deposit parameter within the dataset, batch or plate. '''
    
        pkeysDict, mkeysDict, skeysDict, platesMap, sData, mData, batches, numbatches = args 
        
        #Define figure and subplots
        #plt.close('all')
        if plotType in ['plates', 'batches']:
            fig = plt.figure('%s' %stockinput, figsize=plt.figaspect(0.4))
            kwargs = {'projection':'3d'}
            numplots= 3
            axes = [fig.add_subplot(1, 3, i+1, **kwargs) for i in xrange(numplots)]
            ax1, ax2, ax3 = axes
        elif plotType == 'dataset':
            fig = plt.figure('%s' %stockinput, figsize=plt.figaspect(0.65))
            fig.suptitle('%s: deposits %s distributions' %(stockinput, parameter), fontsize=14) 
            panel1 = [plt.subplot2grid((3,3), (i,0), colspan=1) for i in xrange(3)]
            ax4 = plt.subplot2grid((3,3), (0,1), rowspan = 3, colspan =3, projection = '3d')
            ax1, ax2, ax3 = panel1
            axes = [ax1, ax2, ax3, ax4]
        
        #Define xdata
        xranges = {'size': [420.0], 'hue': [360.0], 'lightness': [1.0], 'saturation': [1.0]}
        maxval= xranges[parameter][0]
        minval = float(minval)
        binsize = maxval/float(numbins)
        xdset = [i for i in np.arange(0.0, maxval, binsize)]
        #Fetch ydata
        metricsDict = {'size': ['seqA', 'size', 'errSize'], 'hue': ['seqH', 'hue', 'errH'], 'lightness': ['seqL', 'lightness', 'errL'], 'saturation': ['seqS', 'saturation', 'errS']}
        skey, mkey, errkey = metricsDict[parameter]
        seqD = sData[skeysDict[skey]]
        meanSeq = mData[mkeysDict[mkey]]
        errSeq = mData[mkeysDict[errkey]]
        
        #Bin ydata: plates
        binnedPlates = [[[np.histogram(plate, range = (0.0, maxval), bins = len(xdset), density = False)[0] for plate in batch] for batch in dtype] for i, dtype in enumerate(seqD)]
        if plotType == 'plates':
            batchinput = int(raw_input('\nPlease indicate the number of the batch to plot (%s). ' %batches))
            if batchinput in batches:
                batchToPlot = batches.index(batchinput)
                platesToPlot = [dtype[batchToPlot] for dtype in binnedPlates]
            else: 
                print('Stock %s is not present in batch QAFF%i.' %(stockinput, batchinput))
                sys.exit()
        #Bin ydata: batches
        zippedBatches = [[zip(*batch) for batch in dtype] for dtype in binnedPlates]
        binnedBatches = [[[np.mean(zipval) for zipval in chain(batch)] for batch in dtype] for dtype in zippedBatches]
        err_binnedBatches = [[[sem(zipval) for zipval in chain(batch)] for batch in dtype] for dtype in zippedBatches]
        #Bin ydata: deposit populations
        zippedDtype = [zip(*dtype) for dtype in binnedBatches]
        binnedDtype = [[np.mean(zipval) for zipval in chain(dtype)] for dtype in zippedDtype]
        zipped_errBinnedBatches = [zip(*dtype) for dtype in err_binnedBatches]
        err_binnedDtype = [[sqrt(sum(np.power(zipval,2))) for zipval in chain(dtype)] for dtype in zipped_errBinnedBatches]
        
        #Define PolyCollection objects: vertices containers
        if plotType == 'batches':
            verts = [[zip(xdset, batch) for batch in dtype] for dtype in binnedBatches]
        elif plotType == 'plates':
            verts = [[zip(xdset, batch) for batch in dtype] for dtype in platesToPlot]
            print(verts[:]); sys.exit()
        #Define PolyCollection objects: generate colorMaps
        cc = lambda arg: colorConverter.to_rgba(arg, alpha=0.6)
        if plotType == 'batches':
            step = 1/float(len(batches))
            hsvColorMap = hsvGenerator(step, 0.8, 0.8)
            rgbaColorMap = [cc(hsv_to_rgb(h,s,v)) for h,s,v in hsvColorMap]
        elif plotType == 'plates':
            batchSizes = [len(dtype) for dtype in platesToPlot]
            steps = [1/float(item) for item in batchSizes]
            hsvColorMaps = [hsvGenerator(step, 0.8, 0.8) for step in steps]
            rgbaColorMap = [[cc(hsv_to_rgb(h,s,v)) for h,s,v in hsvColorMap] for hsvColorMap in hsvColorMaps]
        elif plotType == 'dataset':
            step = 1/float(3)
            hsvColorMap = hsvGenerator(step, 0.8, 0.8)
            rgbaColorMap = [cc(hsv_to_rgb(h,s,v)) for h,s,v in hsvColorMap]
        #Define PolyCollection objects: collection containers
        if plotType in ['plates', 'batches']:
            if plotType == 'batches':
                poly = [PolyCollection(sublist, facecolors = rgbaColorMap) for sublist in verts]
            elif plotType == 'plates':
                poly = [PolyCollection(sublist, facecolors = rgbaColorMap[i]) for i, sublist in enumerate(verts)]
            [sublist.set_alpha(0.7) for sublist in poly]       
        #Add PolyCollection containers to subplots
        if plotType == 'batches':
            zs = np.arange(numbatches)
            [ax.add_collection3d(poly[i], zs = zs, zdir = 'y') for i, ax in enumerate(axes)]
        elif plotType == 'plates':
            zs = [np.arange(len(dtype)) for dtype in platesToPlot]
            [ax.add_collection3d(poly[i], zs = zs[i], zdir = 'y') for i, ax in enumerate(axes)]
        elif plotType == 'dataset':
            bar_width = maxval/(float(numbins) + 10)
            error = [[[0]*len(err), err] for err in err_binnedDtype]
            [ax.bar(xdset, binnedDtype[i], width = bar_width, alpha = 0.3 , color=rgbaColorMap[i],  yerr = error[i],  ecolor = '#BEC4C3', label = ('%s' %stockinput)) for i, ax in enumerate(axes[:-1])]
            [ax.plot(xdset, binnedDtype[i], color = 'g', linewidth = 2) for i, ax in enumerate(axes[:-1])]
            for z in xrange(len(binnedDtype)):
                cs = [rgbaColorMap[z]] * len(xdset)
                ax4.bar(xdset, binnedDtype[z], zs = z, zdir = 'y', width = bar_width, color = cs)
        
        #Set axis limits
        if plotType in ['plates', 'batches']:
            if plotType =='batches':
                [ax.set_ylim3d([0, len(batches)]) for ax in axes]
                zlim = [max([max(batch) for batch in dtype]) for dtype in binnedBatches]
            elif plotType == 'plates':
                [ax.set_ylim3d([0, batchSizes[i]]) for i, ax in enumerate(axes)]
                zlim = [max([max(plate) for plate in dtype]) for dtype in platesToPlot]
            zlim = [int(val) + val/20 for val in zlim]
            [ax.set_zlim3d([0, zlim[i]]) for i, ax in enumerate(axes)]
            [ax.set_xlim3d([0, maxval]) for ax in axes]         
        elif plotType == 'dataset':
            [ax.set_xlim([0, maxval]) for ax in axes]
            ylim = [max(dtype) for dtype in binnedDtype]
            ylim = [int(val) + val/2 for val in ylim]
            [ax.set_ylim([0, ylim[i]]) for i, ax in enumerate(axes[:-1])]
            ax4.set_ylim3d([0, 3])
            ax4.set_zlim3d([0, max(ylim)+10]) 
        #Set axis labels
        if plotType in ['plates', 'batches']:
            [ax.set_xlabel('Deposit %s (pixel)' %parameter, fontsize = 10) for ax in axes]
            [ax.set_zlabel('Deposit count', fontsize = 10) for ax in axes]
        elif plotType == 'dataset':
            ax3.set_xlabel('Deposit %s (pixel)' %parameter, fontsize = 11) 
            ax4.set_xlabel('Deposit %s (pixel)' %parameter, fontsize = 11)
            ax4.set_zlabel('Deposit count', fontsize = 11)
            #ax1.set_ylabel('Deposit count (fly-1 * plate-1)', fontsize = 12) 
        #Set axes titles
        if plotType in ['plates', 'batches', 'dataset']:
            titles = ['RODs', 'nRODs: females', 'nRODs: males']
            if plotType in ['plates', 'batches']:
                [ax.set_title('%s' %titles[i], fontsize = 11) for i, ax in enumerate(axes)]
            elif plotType == 'dataset':    
                [ax.set_title('%s' %titles[i], fontsize = 11) for i, ax in enumerate(axes[:-1])]
        #Define yticklabels
        if plotType in ['plates', 'batches']:
            if plotType == 'batches':
                yticklabels = ['QAFF%i' %batch for batch in batches]
                for ax in axes:
                    ax.set_yticks(zs)
                    ax.set_yticklabels(yticklabels, rotation = -20, verticalalignment = 'baseline', horizontalalignment = 'left', fontsize = 10)     
            elif plotType == 'plates':
                genderKeys = ['females', 'females', 'males']
                plateIDs = [[plate for plate in pkeysDict[batchinput][key]] for key in genderKeys]
                #Convert plateIDs to scanids
                scanids = [[platesMap[label]['qkey'] for label in sublist] for sublist in plateIDs]
                #Set ytick labels
                for i, ax in enumerate(axes):
                    ax.set_yticks(zs[i])
                    ax.set_yticklabels(scanids[i], rotation = -20, verticalalignment = 'baseline', horizontalalignment = 'left', fontsize = 9)       
            [ax.tick_params(axis = 'x', labelsize = 9) for ax in axes]#set xtick labels fontsize            
            [ax.tick_params(axis = 'z', labelsize = 9) for ax in axes]#set ztick labels fontsize   
        elif plotType == 'dataset':
            ax4.set_yticks(np.arange(len(titles)))
            ax4.set_yticklabels(titles, rotation = -10, verticalalignment = 'baseline', horizontalalignment = 'left', fontsize = 12)
            axis = ['x', 'y', 'z']
            [[ax.tick_params(axis = item, labelsize = 9) for item in axis] for ax in axes[:]]
            plt.figtext(0.011, 0.5, 'Deposit count (fly-1 * plate-1)', fontsize = 11, rotation = 'vertical', horizontalalignment='center',verticalalignment='center', transform=ax.transAxes)               
        plt.tight_layout()
        plt.show() 
        return
    
    def batchPlotter(self, parameter, dtype = 'females'):
        from matplotlib import pyplot as plt
        import seaborn as sns
        from itertools import chain               
        #fetch data objects
        metricsDict = {'size': 'seqA', 'hue': 'seqH', 'lightness': 'seqL', 'saturation': 'seqS', 'count': 'normCount', 'rodcount': 'rodFraction'}                          
        batchmetrics = self.loadBatchMetrics()
        batchdata = batchmetrics[metricsDict[parameter]][dtype]                                                   
        #define some data variables
        stocksPerBatch = [len(batch) for batch in batchdata]
        dsetbatches = [i for i in xrange(22)]
        batchmeans = [np.mean(batch) for batch in batchdata]
        #x and y data
        xdata = [[i+1]*stocknumb for i, stocknumb in enumerate(stocksPerBatch)]
        xdata = list(chain(*xdata))
        ydata = list(chain(*batchdata))
        #define figure settings 
        plt.close('all')
        sns.set_style("whitegrid")
        ax = plt.subplot(111)
        #define plots
        if metricsDict[parameter].startswith('seq'):
            sns.boxplot(x = xdata, y = ydata, ax = ax)
        else:
            sns.swarmplot(x = xdata, y = ydata, s = 3, hue = xdata, ax = ax, zorder = -1)
        ax.scatter(dsetbatches, batchmeans, color = 'b', marker = '*', s = 30, label = 'Batch mean', zorder = 1)#plot batch means
        #set xtick labels
        ax.set_xticklabels(['QAFF%s' %str(i+1) for i in xrange(22)], rotation = 60, size = 10)
        #set ylabels and limits
        if parameter == 'count':
            ax.set_ylabel('Number of deposits per fly', fontsize = 13)
        elif parameter == 'rodcount':
            ax.set_ylabel('Fraction of RODs per fly', fontsize = 13)    
        elif parameter in ['lightness', 'saturation', 'hue', 'size']:
            ax.set_ylabel('Deposit %s' %parameter, fontsize = 13)
        #set ylimit
        if parameter in ['count', 'rodcount']:
            maxim = max(ydata)
            ax.set_ylim([0, maxim + maxim/10.0])
        elif parameter in ['lightness', 'saturation']:
            ax.set_ylim([0, 1.0])
        #define legend
        ax.legend([])    
        #handles, labels = ax.get_legend_handles_labels()
        #ax.legend(handles[-1:], labels[-1:], fontsize = 12, loc = 'best')
        plt.show()
        return
    
    
    
class RODFiltering(DataOperations):
    
    def __init__(self):
        DataOperations.__init__(self)
        self.dropimagedir = os.path.join(self.workdir, 'DropImage')
        if not os.path.exists(self.dropimagedir):
                os.makedirs(self.dropimagedir)
        self.sfilterSpreads = self.loadSizeFilterSpreads()
                
    def loadSizeFilterSpreads(self):
        sfilterSpreads =dict(zip(self.batchnumbers, [6]*10 + [4]*12))
        return sfilterSpreads
            
    def generateSLFilterSeq(self, batch, step = 5, maxval = [1.0, 0.6]):
        from itertools import ifilter
        from File_Functions import listPartition
        '''It filters RODs for a given batch using a sequence of s and l values, generated from user-defined max values 
        for s and l (maxval[1] and maxval[0]); step defines the spacing between consecutive s and l values. '''
        #Load dictionaries
        depositsDict = self.depositsDict(batch)
        platesDict = self.platesDict(batch)
        platesMap = self.loadPlatesMap(batch = batch)
        #Filter out headings from keyset
        keyset = [key for key in depositsDict.keys() if key not in self.dheads]
        #Fetch ROD positive depositkeys
        dkeysROD = [key for key in keyset if depositsDict[key]['ROD'] == 1]
        #Fetch plate keys from rod_depositkeys 
        pkeysROD = list(set([int(key.split('_')[1]) for key in dkeysROD]))
        #Sort pKeys according to gender
        pkeys = sorted([int(key.split('_')[1]) for key in platesMap.keys()])
        pkeysROD = [key for key in pkeysROD if key in pkeys]
        pkeysROD_cluster = listPartition(lambda x:platesMap['%s_%s'%(batch, x)]['stockId'][-2] == 'F', pkeysROD)
        ##Remap deposit keys to pKeys clusters
        dkeysROD_cluster = [[key for key in dkeysROD if int(key.split('_')[1]) in gender_sublist] for gender_sublist in pkeysROD_cluster]
        #Generate rodkeys sets
        pkeysROD_set = [list(set([int(key.split('_')[1]) for key in gender_sublist])) for gender_sublist in dkeysROD_cluster]
        [sorted(subset) for subset in pkeysROD_set]
        #Fetch lightness and saturation values
        rodLS = [[(key, depositsDict[key]['meanL'], depositsDict[key]['meanS']) for key in gender_sublist] for gender_sublist in dkeysROD_cluster]
        #Define saturation and ligthness filter sequences
        filter_s = np.linspace(maxval[1], 0.0, num = step)    
        filter_l = np.linspace(maxval[0], 0, num = step)
        #Filter out putative ROD false positives
        rodLS_filtered = []
        #Filter for saturation 
        for sublist in rodLS: 
            sublist = sorted(sublist, key = lambda x:int(x[0].split('_')[0])) #Sort gender_sublists 
            supralist = [[i for i in ifilter(lambda x:x[2]<=filter_s[0], sublist)]] 
            for val in filter_s[1:]:
                supralist.append([i for i in ifilter(lambda x:x[2]<=val, supralist[-1])])#Filter sequentially
        
            rodLS_filtered.append(supralist)
            supralist = []
        #Filter for lightness
        for i in xrange(len(rodLS_filtered)):  
            for j, sublist in enumerate(rodLS_filtered[i]):
                supralist = [[item for item in ifilter(lambda x:x[1]<=filter_l[0], sublist)]]
                for val in filter_l[1:]:
                    supralist.append([item for item in ifilter(lambda x:x[1]<=val, supralist[-1])])
            
                rodLS_filtered[i][j] = supralist
                supralist = []
        #Fetch sqlkeys from filtered sublists
        rodkeys_filtered = [[[[tupl[0].split('_')[1] for tupl in sublist] for sublist in sublist_s] for sublist_s in gender_sublist] for gender_sublist in rodLS_filtered]
        #Fetch number of RODs after filtering
        rodcount_filtered = [[[[sublist.count(str(key)) for key in pkeysROD_set[i]] for sublist in sublist_s] for sublist_s in gender_sublist] for i, gender_sublist in enumerate(rodkeys_filtered)]
        #Fetch number of RODs before filtering
        rodcount = [[platesDict[key]['numberRODs'] for key in sublist] for sublist in pkeysROD_cluster]
        #pack data
        data = [rodcount, rodcount_filtered, filter_s, filter_l]
        return data
        
    def optimizeRODcFilter(self, batch, step = 10, maxval = [1.0, 0.6]):
        from itertools import chain
        import numpy as np
        #fetch data from ROD filter
        data = self.generateSLFilterSeq(batch, step = step, maxval = maxval)
        #unpack data
        rodcount, rodcount_filtered, filter_s, filter_l = data
        countF, countM = rodcount
        countF_filtered, countM_filtered = rodcount_filtered
        #count number of RODs for each sl filter pair
        mlist = [[np.sum(llist) for llist in slist] for slist in countM_filtered]
        flist = [[np.sum(llist) for llist in slist] for slist in countF_filtered]
        #compare male and female ROD counts for the same sl filter pair - difference
        slarray = [[(val,(i,j), flist[i][j]-val) for j, val in enumerate(sublist)] for i, sublist in enumerate(mlist)]
        #sort tuples and filter out sl filter pairs that 
        slarray = sorted(list(chain(*slarray)), key = lambda x:x[0])
        slarray = [val for val in slarray if val[2]>0]
        #fetch optimal sl filter pair
        maxpossible = max([c for (a,b,c) in slarray])#fetch maximum possible ROD count
        slarray = [(a,b, maxpossible-c) for (a,b,c) in slarray]#reformat slarray tuples 
        slarray = sorted([(np.sqrt(np.power(a,2)+np.power(c,2)), b) for (a,b,c) in slarray], key = lambda x:x[0])[:1]#compute distance of (a,c) to (0, maxpossible) and fetch closest point
        sltuple = [(filter_s[b[0]], filter_l[b[1]]) for (a,b) in slarray]#fetch sl parameters
        return sltuple
        
            
    def fecthRODPopulation(self, batchlim = [1,10]):
        #load data objects
        lsDict = self.loadRODColourFilter()
        batches = [batch for batch in lsDict.keys()  if batchlim[0] <= batch <= batchlim[1]]
        #fetch ROD dataset
        rodsize_dset = []    
        for batch in batches:
            s,l = lsDict[batch]
            print('QAFF%s' %batch, l, s)
            #fetch deposits object and RODs data
            rodHLS = self.fetchRODsData(batch)
            depositsDict = self.depositsDict(batch)
            #filter out putative ROD false positives
            rod_filtered = [[tupl for tupl in sublist if tupl[2] <= l and tupl[3] <= s] for sublist in rodHLS] 
            rodkeys_filtered = [[tupl[0] for tupl in sublist] for sublist in rod_filtered]
            rodkeys_filtered = [sorted(sublist, key = lambda x: (int(x.split('_')[1]), int(x.split('_')[0]))) for sublist in rodkeys_filtered]
            #Fetch deposits size metrics
            rodsize_filtered = [[(depositsDict[key]['area'], depositsDict[key]['perimeter']) for key in sublist] for sublist in rodkeys_filtered]
            rodsize = [zip(*sublist) for sublist in rodsize_filtered]
            rodsize_dset.append(rodsize)       
        return rodsize_dset
    
    
    def plotSizeFilterROI(self, data = 'females', batchlim = [1,10], spread = 6.0):
        import matplotlib.pyplot as plt
        from matplotlib.patches import Ellipse
        from statsFunctions import pointInEllipse
        from itertools import chain
        #fetch data objects
        print('Fetching ROD population: QAFF%s to QAFF%s' %(batchlim[0], batchlim[1]))
        rodsize_dset = self.fecthRODPopulation(batchlim = batchlim)
        sizeFilter = self.loadRODSizeFilter()
        #unpack data
        rodsizeF, rodsizeM = zip(*rodsize_dset)
        rodsizeF =zip(*rodsizeF)
        rodsizeM =zip(*rodsizeM)
        areaF, perimeterF = [np.asarray(list(chain(*sublist))) for sublist in rodsizeF]#reformat as numpy array
        areaM, perimeterM = [np.asarray(list(chain(*sublist))) for sublist in rodsizeM]#reformat as numpy array
        areas = [areaF, areaM]
        perimeters = [perimeterF, perimeterM]
        #apply box cox transformation to data
        print('Transforming data: boxcox')
        (xlambda_cox, ylambda_cox) = sizeFilter['boxcox']#boxcox lambda factors
        xcox = [(gender**xlambda_cox - 1) / xlambda_cox  for gender in areas]
        ycox = [(gender**ylambda_cox - 1) / ylambda_cox for gender in perimeters]
        #fetch normal bivariate fits
        print('Estimating bivariate normal fit')
        (mu_nr, sigma1_nr,sigma2_nr, alpha_nr) = sizeFilter['nrob']
        (mu_r, sigma1_r,sigma2_r, alpha_r) = sizeFilter['rob']
        #test whether points are inside elliptic surface
        colors = [['b' if pointInEllipse(mu_nr[0], mu_nr[1], valx, ycox[i][j], sigma1_nr*spread, sigma2_nr*spread, alpha_nr) else '#E1E4E6' for j, valx in enumerate(gender)] for i, gender in enumerate(xcox)]
        #scatter the points
        ax = plt.subplot(111)
        datakeys = ['females', 'males']; idx = datakeys.index(data)
        ax.scatter(xcox[idx], ycox[idx], s = 8, c = colors[idx], alpha = 0.3)
        #ax.scatter(areaF, perimeterF)
        # Draw elipses showing the fits
        for Nsig in [spread]:
            #Non-robust fit
            E_nr = Ellipse(mu_nr, sigma1_nr * Nsig, sigma2_nr * Nsig,
                    (alpha_r * 180. / np.pi),
                        ec='g', fc='none', linestyle='solid', lw=2.0)
            ax.add_patch(E_nr)
            #Robust fit
            E_r = Ellipse(mu_r, sigma1_r * Nsig, sigma2_r * Nsig,
                        (alpha_r * 180. / np.pi),
                        ec='r', fc='none', linestyle='dashed', lw = 2.5)
            ax.add_patch(E_r)
        plt.show()
        return

                        
    def plotSLFilterSeq(self, batch, step = 5, maxval = [1.0, 0.6]):
        from matplotlib import pyplot as plt
        #plt.switch_backend('WXAgg')        
        from Plot_Functions import gridSize
        '''It takes a batch number as input and returns various graphs of the ROD count for each gender, before and after filtering. RODs 
        are filtered using a sequence of s and l values, generated from user-defined max values for s and l (maxval[1] and maxval[0]); step 
        defines the spacing between consecutive s and l values. Each graph plots the ROD count for a specific s value and a sequence of l values. '''
        #fetch data from ROD filter
        data = self.generateSLFilterSeq(batch, step = step, maxval = maxval)
        #unpack data
        rodcount, rodcount_filtered, filter_s, filter_l = data
        countF, countM = rodcount
        countF_filtered, countM_filtered = rodcount_filtered
        #Plot RODs counts
        plt.close('all')
        plt.figure('QAFF%i: RODs analysis' %batch)
        #Define grid size
        rows, cols = gridSize(2*step)
        #Define axes positions in the grid
        axlist = [ax for ax in xrange(rows*cols)]
        count = 0
        for i in xrange(rows):
            for j in xrange(cols):
                axlist[count] = plt.subplot2grid((rows,cols), (i,j), colspan=1)
                count +=1       
        #Females
        xdatasetF = np.arange(len(countF))
        for i, slist in enumerate(countF_filtered):
            for j in xrange(step):
                axlist[i].plot(xdatasetF, slist[j],  color = 'b', linewidth = 0.2)
            axlist[i].plot(xdatasetF, countF, color = 'r',  linewidth = 0.4) 
        #Males
        xdatasetM = np.arange(len(countM))    
        for i, slist in enumerate(countM_filtered):
            for j in xrange(step):
                axlist[(cols-1)+i].plot(xdatasetM, slist[j],  color = 'b', linewidth = 0.2)
            axlist[cols+i].plot(xdatasetM, countM, color = 'r',  linewidth = 0.4)     
        #Set axis labels and titles
        for i, ax in enumerate(axlist):
            try:
                if i <= cols-1:
                    ax.set_title('Females: S<=%.2f' %filter_s[i], fontsize = 12)
                    if i == 0:
                        ax.set_ylabel('RODs count')
                elif i >= cols:
                    ax.set_title('Males: S<=%.2f' %filter_s[i-cols], fontsize = 12)
                    ax.set_xlabel('PlateID')
                    if i == cols:
                        ax.set_ylabel('RODs count')            
            except IndexError:
                continue           
        plt.tight_layout()
        plt.show()
        return
    
    
    def fetchRODsData(self, batch):
        from File_Functions import listPartition
        '''It takes a batch number as input and fetches for each ROD deposit its metrics. It returns a list of named tuples
        [(key,hue), (key,lightness), (key, saturation)] for each deposit, clustered according to plate.'''
        #Load dictionaries
        depositsDict = self.depositsDict(batch)
        platesMap = self.loadPlatesMap(batch = batch)
        #Filter out headings from keyset
        keyset = [key for key in depositsDict.keys() if key not in self.dheads]
        #Fetch ROD positive sqlkeys 
        dkeysROD = [key for key in keyset if depositsDict[key]['ROD'] == 1]
        #Fetch plate keys from rod_depositkeys 
        pkeysROD = list(set([int(key.split('_')[1]) for key in dkeysROD]))
        #Sort pKeys according to gender
        pkeys = sorted([int(key.split('_')[1]) for key in platesMap.keys()])
        pkeysROD = [key for key in pkeysROD if key in pkeys]
        pkeysROD_cluster = listPartition(lambda x:platesMap['%s_%s'%(batch, x)]['stockId'][-2] == 'F', pkeysROD)
        ##Remap deposit keys to pKeys clusters
        dkeysROD_cluster = [[key for key in dkeysROD if int(key.split('_')[1]) in gender_sublist] for gender_sublist in pkeysROD_cluster]
        #Fetch hue and saturation values
        rodHLS = [[(key, depositsDict[key]['meanH'], depositsDict[key]['meanL'], depositsDict[key]['meanS']) for key in sublist] for sublist in dkeysROD_cluster]
        return rodHLS
        
    
    def filterRODs(self, batch, spread = 3.0):
        '''It takes a batch number as input and filters RODs keys and scan coordinates using optimised s and l values. 
        It returns a list of RODs filtered data: RODs keys, RODs xy coordinates clustered according plate source and ROD count after filtering.'''
        from statsFunctions import pointInEllipse
        #fetch s,l values
        lsDict = self.loadRODColourFilter()
        s,l = lsDict[batch]
        #fetch deposits object and RODs data
        rodHLS = self.fetchRODsData(batch)
        depositsDict = self.depositsDict(batch)
        #filter out putative ROD false positives by color
        rod_filtered = [[tupl for tupl in sublist if tupl[2] <= l and tupl[3] <= s] for sublist in rodHLS] 
        rodkeys_filtered = [[tupl[0] for tupl in sublist] for sublist in rod_filtered]
        rodkeys_filtered = [sorted(sublist, key = lambda x: (int(x.split('_')[1]), int(x.split('_')[0]))) for sublist in rodkeys_filtered]
        #Fetch RODs size data
        rodsize_filtered = [[(depositsDict[key]['area'], depositsDict[key]['perimeter']) for key in sublist] for sublist in rodkeys_filtered]
        #unpack data
        rodsize_filtered = [zip(*sublist) for sublist in rodsize_filtered]
        rodsizeF, rodsizeM = rodsize_filtered
        areaF, perimeterF = [np.asarray(sublist) for sublist in rodsizeF]#reformat as numpy array
        areaM, perimeterM = [np.asarray(sublist) for sublist in rodsizeM]#reformat as numpy array
        areas = [areaF, areaM]
        perimeters = [perimeterF, perimeterM]
        #filter out putative ROD false positives by size
        #apply box cox transformation to data
        xlambda_cox = -0.31451158063953116; ylambda_cox = -0.80645096067649358#boxcox lambda factors
        xcox = [(gender**xlambda_cox - 1) / xlambda_cox  for gender in areas]
        ycox = [(gender**ylambda_cox - 1) / ylambda_cox for gender in perimeters]
        #load RODS sizeFilter
        sizeFilter = self.loadRODSizeFilter()
        #fetch normal bivariate fits
        (mu_nr, sigma1_nr,sigma2_nr, alpha_nr) = sizeFilter['nrob']
        #test whether points are inside elliptic surface
        sizefiltered = [[pointInEllipse(mu_nr[0], mu_nr[1], valx, ycox[i][j], sigma1_nr*spread, sigma2_nr*spread, alpha_nr) for j, valx in enumerate(gender)] for i, gender in enumerate(xcox)]
        sizefiltered = [np.asarray(gender) for gender in sizefiltered]
        rodkeys_filtered = [np.asarray(gender) for gender in rodkeys_filtered]
        rodkeys_2xfiltered = [gender[sizefiltered[i]] for i, gender in enumerate(rodkeys_filtered)]
        #Fetch deposits xy coordinates
        pkeysROD_set = [list(set([int(key.split('_')[1]) for key in sublist])) for sublist in rodkeys_2xfiltered]
        pkeysROD_set = [sorted(sublist) for sublist in pkeysROD_set]
        rodxy_filtered = [[(int(key.split('_')[1]), depositsDict[key]['Xposition'], depositsDict[key]['Yposition']) for key in sublist] for sublist in rodkeys_2xfiltered]
        #Cluster coordinates according to plate source
        rodxy_clusters = [[[(tupl[1], tupl[2]) for tupl in sublist if tupl[0] == key] for key in pkeysROD_set[i]] for i, sublist in enumerate(rodxy_filtered)]
        #Fetch number of RODs after filtering
        rodcount_filtered = [[len(item) for item in sublist] for sublist in rodxy_clusters]
        countF_filtered, countM_filtered = rodcount_filtered
        #pack filtered data
        RODs_filtered = [pkeysROD_set, rodxy_clusters, rodcount_filtered]
        return RODs_filtered
    
    def fetchTypeIRODkeys(self):
        from File_Functions import listPartition, flattenList
        from itertools import ifilter
        #define variables
        typeIRODkeys = {}
        batches = self.batchnumbers
        #filter rod keys for each batch
        for batch in batches:
            print('Analysing QAFF%s' %batch)
            #Load data objects
            depositsDict = self.depositsDict(batch)
            keysmap = self.loadDatasetKeysMap()
            #Filter out headings from keyset
            keyset = [key for key in depositsDict.keys() if key not in Dashboard().dheads]   
            #Fetch ROD and non-ROD depositkeys
            rodkeys, nrodkeys = listPartition(lambda x: depositsDict[x]['ROD'] == 1, keyset)
            #fetch filtered ROD keys
            rodkeys_filtered = keysmap['rod'][batch].values()
            rodkeys_filtered = flattenList(rodkeys_filtered)
            #filter out false positives
            falseRODs = list(ifilter(lambda x: x not in rodkeys_filtered, rodkeys))
            typeIRODkeys[batch] = falseRODs
        return typeIRODkeys
        
    def calculateFDRforRODs(self):
        #load size filter spreads
        sfilterSpreads = RODFiltering().loadSizeFilterSpreads()
        rodFDR = {}
        for batch in self.batchnumbers:
            print('Analysing QAFF%s' %batch)
            #fetch objects and data
            plateDict = self.platesDict(batch)
            spread = sfilterSpreads[batch]
            #filter RODs
            RODs_filtered = self.filterRODs(batch, spread = spread)
            [pkeysROD_set, rodxy_clusters, rodcount_filtered] = RODs_filtered#unpack
            countF_filtered, countM_filtered = rodcount_filtered#unpack
            #Fetch number of RODs before filtering
            rodcount = [[plateDict[key]['numberRODs'] for key in sublist] for sublist in pkeysROD_set]
            countF, countM = rodcount
            a = np.mean(countF_filtered) 
            b = np.mean(countM_filtered)
            ratio = b/a
            values = (a, b, ratio)
            rodFDR[batch] = values
        return rodFDR
        
    def plotRODsFDR(self):            
        import matplotlib.pyplot as plt
        #load data objects                        
        rodFDR = self.loadFDRforRODs()
        countF, countM, ratio = zip(*rodFDR.values())
        batchnumbers = rodFDR.keys()
        #set plot
        plt.close('all')
        ax1 = plt.subplot(111)
        ax1.plot(batchnumbers, countF,  color = 'b', label = 'countF')
        ax1.plot(batchnumbers, countM, color = 'g', label = 'countM')
        ax1.set_xticks(batchnumbers)
        #set yaxis label
        ax1.set_ylabel('ROD count')
        #set tick labels
        labels = ['QAFF%s' %batch for batch in batchnumbers]
        ax1.set_xticklabels(labels, rotation = -60)
        #twin plot
        ax2 = ax1.twinx()
        ax2.plot(batchnumbers, ratio, color = 'r', label = 'rod FDR')
        ax2.set_ylabel('RODs false discovery rate')
        for tl in ax2.get_yticklabels():
            tl.set_color('r')
        ax1.legend()
        #ax2.legend()
        plt.show()
        return          
  
             
    def RODplotter(self, batch, plottype = None, spread = 3.0):
        from matplotlib import pyplot as plt
        #plt.switch_backend('WXAgg')
        from matplotlib.lines import Line2D
        from mpl_toolkits.mplot3d import Axes3D 
        '''It takes a batch number as input and returns a plot of either the colour distribution of the ROD population 
        or the ROD count, for each gender. Colour distributions can either be plotted in 2D (plottype = 'raw2D') or
        3D ('raw3D'). ROD count (plottype = 'filter') is filtered using user-defined s and l values.'''
        #fetch objects and data
        rodHLS = self.fetchRODsData(batch)
        plateDict = self.platesDict(batch)
        platesMap = self.loadPlatesMap(batch = batch)
        #Unpack data
        rodHLS_F, rodHLS_M = rodHLS
        rawkeys_F, arrH_F, arrL_F, arrS_F= zip(*rodHLS_F)
        rawkeys_M, arrH_M, arrL_M, arrS_M= zip(*rodHLS_M)
        if plottype == 'raw2D':
            #set figure
            figure, ax1 = plt.subplots(1, 1)
            #Set window title
            figure.suptitle('QAFF%i: ROD analysis' %batch, fontsize = 16)
            #set plots
            females = ax1.scatter(arrS_F, arrH_F, s = 7, c = 'b')
            males = ax1.scatter(arrS_M, arrH_M, s = 8, c = '#F26BDE', alpha = 0.6, edgecolor = 'none')
            #set xy labels
            ax1.set_xlabel('Saturation', fontsize = 14)
            ax1.set_ylabel('Hue', fontsize = 14)
            #Add legend
            ax1.legend((females, males), ('Female', 'Male'))
            #Set xy limits
            ax1.set_xlim(0, 1)
            ax1.set_ylim(0, 400)
            plt.tight_layout()
            plt.show()   
        elif plottype == 'raw3D':    
            #set figure
            fig = plt.figure('QAFF%i: RODs analysis' %batch)
            ax = Axes3D(fig)
            #set plots
            ax.scatter(arrS_F, arrH_F, arrL_F, s = 7, c = 'b', label = 'Females')
            ax.scatter(arrS_M, arrH_M, arrL_M, s = 8, c = '#F26BDE', alpha = 0.6, edgecolor = 'none', label = 'Males')
            #Add legend
            females = Line2D([0],[0], linestyle="none", c= 'b', marker = 'o')
            males = Line2D([0],[0], linestyle="none", c= '#F26BDE', marker = 'o')
            ax.legend([females, males], ['Females', 'Males'], numpoints = 1, loc = 'best')
            #Set xyz labels
            ax.set_xlabel('Saturation', fontsize = 14)
            ax.set_ylabel('Hue', fontsize = 14)
            ax.set_zlabel('Lightness', fontsize = 14)
            #Set xyz limits
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 375)
            ax.set_zlim(0, 1)
            plt.show()           
        elif plottype == 'filter':
            #fetch data objects
            lsDict = self.loadRODColourFilter()
            s, l = lsDict[batch]
            #filter RODs
            RODs_filtered = self.filterRODs(batch, spread = spread)
            [pkeysROD_set, rodxy_clusters, rodcount_filtered] = RODs_filtered#unpack
            countF_filtered, countM_filtered = rodcount_filtered#unpack
            #Fetch number of RODs before filtering
            rodcount = [[plateDict[key]['numberRODs'] for key in sublist] for sublist in pkeysROD_set]
            countF, countM = rodcount            
            #set figure
            plt.close('all')
            figure, (ax1, ax2) = plt.subplots(2, 1)
            #set window title
            figure.canvas.set_window_title('QAFF%i: RODs analysis' %batch)
            #set xdata
            xdataset_F = np.arange(len(countF))
            xdataset_M = np.arange(len(countM))
            #set axes
            ax1.plot(xdataset_F, countF_filtered, c = 'b', label = 'L<=%.2f, S<=%.2f, sigm = %.1f' %(l, s, spread))
            ax1.plot(xdataset_F, countF, c = '#F26BDE', label = 'No filter')
            ax2.plot(xdataset_M, countM_filtered, c = 'b', label = 'L<=%.2f, S<=%.2f, sigm = %.1f' %(l, s, spread))
            ax2.plot(xdataset_M, countM, c = '#F26BDE', label = 'No filter')
            #set axis ticks and ticks labels
            ax1.set_xticks(xdataset_F)
            ax2.set_xticks(xdataset_M)
            xlabelsF, xlabelsM = [[platesMap['%s_%s'%(batch,key)]['qkey'] for key in sublist] for sublist in pkeysROD_set]
            ax1.set_xticklabels(xlabelsF, rotation = 60, fontsize = 8)
            ax2.set_xticklabels(xlabelsM, rotation = 60, fontsize = 8)
            #set axis labels and axes titles
            ax1.set_title('Females', fontsize = 14)
            ax1.set_ylabel('ROD count', fontsize = 12)
            ax2.set_title('Males', fontsize = 14)
            ax2.set_ylabel('ROD count', fontsize = 12)
            ax2.set_xlabel('Plate', fontsize = 12)
            #set legend
            ax1.legend(loc = 'best', fontsize = 10)
            ax2.legend(loc = 'best', fontsize = 10)
            #filepath = os.path.join(self.dropimagedir, 'RODfiltering_QAFF%s' %batch)
            #figure.savefig(filepath)
            plt.tight_layout()
            plt.show() 
        return
                  
    
    def remarkScans(self, batch, spread = 3.0):
        sys.path.append('C:\Python27\Lib\site-packages')
        import cv2 
        '''It takes a batch number as input and re-marks all plate scans within that batch to reflect ROD filtering.'''
        #load plates dictionary
        platesDict = self.platesDict(batch)
        lsDict = self.loadRODColourFilter()
        s, l = lsDict[batch]
        #filter RODs
        RODs_filtered = self.filterRODs(batch, spread = spread )   
        [pkeysROD_set, rodxy_clusters, rodcount_filtered] = RODs_filtered#unpack   
        #Fetch filenames of annotated images
        rod_scanids = [[platesDict[key]['annotatedImageFileName'] for key in sublist] for sublist in pkeysROD_set]
        #fetch image paths
        annotations_dir = 'U:\QAFF\QAFF%i_scans\Annotated_scans' %batch
        imgpath = [[os.path.join(annotations_dir, filename) for filename in sublist] for sublist in rod_scanids]
        #annotate scans and save annotated images
        output_path = os.path.join(annotations_dir, 'rodFiltered') #Destination directory
        if not os.path.isdir(output_path):
            os.mkdir(output_path)
        for i, subset in enumerate(imgpath):
            for j, path in enumerate(subset):
                img = cv2.imread(path, 1)
                imgname = '%s.png' %os.path.split(path)[1][:-4]
                outpath = os.path.join(output_path,imgname)
                for xy in rodxy_clusters[i][j]:
                    xy = tuple([int(val) for val in xy])
                    cv2.circle(img, xy, 10, (0, 255, 0), thickness=2, lineType=8, shift=0)
                cv2.imwrite(outpath,img)            
        return

        
         
#Dashboard().resetWorkEnv()                                
#DataOperations().purgePickleDir()
#lsDict = QAFFObjects().loadRODColourFilter()
#QAFFObjects().buildDatasetMetricsObj()
#keysmap = QAFFObjects().loadDatasetKeysMap()
#print(keysmap['rod'][1].keys())        
#datasetMetrics = QAFFObjects().loadDatasetMetrics()
#print(datasetMetrics['seq']['normCount']['Empty'])
#platesmap = QAFFObjects().loadPlatesMap()
#workdset = QAFFObjects().loadWorkDataset()
#platesmap = QAFFObjects().loadPlatesMap()
#pdv = PooDataVis()
#stockinput, parameter, plotType = ['js10', 'hue', 'plates']
#pdv.stockPlotter(stockinput, parameter, plotType)
#slDict = QAFFObjects().loadRODColourFilter()        
#RODFiltering().plotSLFilterSeq(8, step = 5, maxval = [1.0, 0.6])
#sltuple = RODFiltering().optimizeRODcFilter(1, step = 10, maxval = [1.0, 0.6])
#DataOperations().calculateMetricsForDataset()
#stockdata = DataOperations().stockDataFetcher('GFPi')
#stockdata = DataOperations().calculateStockMetrics('GFPi')
#platesmap = QAFFObjects().loadPlatesMap()
#DataOperations().calculateStockMetrics('LK')
#RODFiltering().remarkScans(22, spread = 4.0)
#RODFiltering().plotSizeFilterROI(data = 'females', batchlim = [1,10], spread = 6.0)
#RODFiltering().RODplotter(4, plottype = 'filter', spread = 6.0)
#rodDset = RODFiltering().fecthRODPopulation(batchlim = [1,10])



